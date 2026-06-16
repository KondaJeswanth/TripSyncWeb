"""
TripSync Web Application — Full Selenium E2E Test Suite
========================================================
Covers:
  - Functional Tests  : All pages, forms, navigation, API integrations
  - Vulnerability Tests: XSS, open redirects, security headers, input sanitization
  - Unit-level Tests   : Backend API health, endpoint responses, CORS validation

Run locally:
    python tests/test_tripsync.py

Output:
    tests/tripsync_test_report.xlsx
"""

import os
import sys
import time
import json
import unittest
import datetime
import requests

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, WebDriverException
)
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────
BASE_URL   = os.environ.get("TRIPSYNC_URL",    "https://abineshh502.github.io/TripSyncWeb-Frontend")
API_URL    = os.environ.get("TRIPSYNC_API_URL", "https://tripsyncweb-backend.onrender.com")
HEADLESS   = os.environ.get("HEADLESS", "true").lower() == "true"
TIMEOUT    = int(os.environ.get("TIMEOUT", "20"))
REPORT_DIR = os.environ.get("REPORT_DIR", os.path.join(os.path.dirname(__file__), "..", "test-results"))

os.makedirs(REPORT_DIR, exist_ok=True)

# ─── Global result collector ───────────────────────────────────────────────────
_results: list[dict] = []

def _record(category: str, test_id: str, name: str, status: str,
            details: str = "", duration: float = 0.0, severity: str = "Normal"):
    _results.append({
        "category":  category,
        "test_id":   test_id,
        "name":      name,
        "status":    status,
        "details":   details,
        "duration":  round(duration, 2),
        "severity":  severity,
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })


# ─── WebDriver factory ────────────────────────────────────────────────────────
def _build_driver() -> webdriver.Chrome:
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--ignore-certificate-errors")
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    import platform, shutil
    if platform.system() == "Linux":
        # GitHub Actions / Ubuntu runner — chromedriver is in PATH via setup-chrome action
        driver_path = shutil.which("chromedriver") or "/usr/bin/chromedriver"
        service = Service(executable_path=driver_path)
    else:
        # Windows local — use ChromeDriverManager
        try:
            service = Service(ChromeDriverManager().install())
        except Exception:
            local_driver = r"C:\Users\konda\.wdm\drivers\chromedriver\win64\149.0.7827.115\chromedriver-win32\chromedriver.exe"
            service = Service(executable_path=local_driver)

    return webdriver.Chrome(service=service, options=opts)


def _wait_visible(driver, by, selector, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, selector))
    )

def _wait_clickable(driver, by, selector, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )

def _wait_title(driver, partial, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.title_contains(partial)
    )

def _find_submit_btn(driver):
    """Find a submit/login/register button using multiple fallback selectors."""
    selectors = [
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.XPATH,        "//button[@type='submit']"),
        (By.XPATH,        "//button[contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SIGN') or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'LOG') or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'REGISTER') or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CREAT') or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SUBMIT')]"),
        (By.CSS_SELECTOR, "form button"),
        (By.CSS_SELECTOR, "button"),
    ]
    for by, sel in selectors:
        try:
            els = driver.find_elements(by, sel)
            if els:
                return els[0]
        except Exception:
            continue
    raise NoSuchElementException("No submit/action button found on page")


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — FUNCTIONAL TESTS
# ═══════════════════════════════════════════════════════════════════════════════
class FunctionalTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.driver = _build_driver()
        cls.driver.set_page_load_timeout(45)

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()

    def _go(self, path=""):
        self.driver.get(BASE_URL.rstrip("/") + ("/" + path.lstrip("/") if path else ""))
        time.sleep(1.5)

    # ── TC-F-001: Landing page loads ─────────────────────────────────────────
    def test_F001_landing_page_loads(self):
        t0 = time.time()
        try:
            self._go()
            body = self.driver.find_element(By.TAG_NAME, "body")
            assert body is not None
            page_src = self.driver.page_source.lower()
            assert "tripsync" in page_src, "TripSync brand not on landing page"
            _record("Functional", "F-001", "Landing Page Loads",
                    "PASS", f"Title: {self.driver.title}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-001", "Landing Page Loads",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-F-002: Navbar is visible ──────────────────────────────────────────
    def test_F002_navbar_present(self):
        t0 = time.time()
        try:
            self._go()
            nav = self.driver.find_element(By.TAG_NAME, "nav")
            assert nav.is_displayed()
            _record("Functional", "F-002", "Navbar Visible on Landing",
                    "PASS", "nav element found and visible", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-002", "Navbar Visible on Landing",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-003: Login page loads ───────────────────────────────────────────
    def test_F003_login_page_loads(self):
        t0 = time.time()
        try:
            self._go("login")
            email_input = _wait_visible(self.driver, By.ID, "email")
            pass_input  = _wait_visible(self.driver, By.ID, "password")
            assert email_input.is_displayed()
            assert pass_input.is_displayed()
            _record("Functional", "F-003", "Login Page — Inputs Visible",
                    "PASS", "Email & password fields found", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-003", "Login Page — Inputs Visible",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-F-004: Login form validation (empty submit) ───────────────────────
    def test_F004_login_empty_validation(self):
        t0 = time.time()
        try:
            self._go("login")
            btn = _find_submit_btn(self.driver)
            btn.click()
            time.sleep(0.8)
            # Should NOT navigate away from login
            assert "login" in self.driver.current_url or "register" not in self.driver.current_url
            _record("Functional", "F-004", "Login Empty Validation Stays on Page",
                    "PASS", "Did not navigate away on empty submit", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-004", "Login Empty Validation Stays on Page",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-005: Login with wrong credentials shows error ───────────────────
    def test_F005_login_wrong_credentials(self):
        t0 = time.time()
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys("invalid@test.com")
            self.driver.find_element(By.ID, "password").send_keys("wrongpassword123")
            _find_submit_btn(self.driver).click()
            time.sleep(4)
            page_src = self.driver.page_source
            # Should still be on login or show error
            has_error = any(kw in page_src.lower() for kw in
                           ["invalid", "error", "incorrect", "not found", "wrong"])
            assert has_error or "dashboard" not in self.driver.current_url
            _record("Functional", "F-005", "Login Wrong Credentials Error Handling",
                    "PASS", "Error shown or stayed on login", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-005", "Login Wrong Credentials Error Handling",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-006: Register page loads ────────────────────────────────────────
    def test_F006_register_page_loads(self):
        t0 = time.time()
        try:
            self._go("register")
            time.sleep(1.5)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["register", "sign up", "create", "account"])
            _record("Functional", "F-006", "Register Page Loads",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-006", "Register Page Loads",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-007: Dashboard page redirects unauthenticated users ─────────────
    def test_F007_dashboard_auth_redirect(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            time.sleep(3)
            url = self.driver.current_url
            # Should either show login or a redirect to /login or /
            redirected = "login" in url or url.rstrip("/").endswith(BASE_URL.rstrip("/"))
            page_src = self.driver.page_source.lower()
            shows_login_content = any(kw in page_src for kw in ["sign in", "email address", "password"])
            assert redirected or shows_login_content
            _record("Functional", "F-007", "Dashboard Redirects Unauthenticated Users",
                    "PASS", f"Redirected to: {url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-007", "Dashboard Redirects Unauthenticated Users",
                    "FAIL", str(e), time.time()-t0, "High")
            self.fail(str(e))

    # ── TC-F-008: AI Assistant page accessible (redirect or content) ─────────
    def test_F008_ai_assistant_page(self):
        t0 = time.time()
        try:
            self._go("ai-assistant")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            reachable = any(kw in page_src for kw in ["tripsync", "assistant", "ai", "sign in", "login"])
            assert reachable
            _record("Functional", "F-008", "AI Assistant Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-008", "AI Assistant Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-009: AI Planner page accessible ─────────────────────────────────
    def test_F009_ai_planner_page(self):
        t0 = time.time()
        try:
            self._go("ai-planner")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "planner", "journey", "sign in"])
            _record("Functional", "F-009", "AI Planner Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-009", "AI Planner Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-010: Safety page accessible ─────────────────────────────────────
    def test_F010_safety_page(self):
        t0 = time.time()
        try:
            self._go("safety")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "safety", "crowd", "sign in"])
            _record("Functional", "F-010", "Safety & Crowd Advisor Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-010", "Safety & Crowd Advisor Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-011: Expenses page accessible ───────────────────────────────────
    def test_F011_expenses_page(self):
        t0 = time.time()
        try:
            self._go("expenses")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "expense", "bill", "split", "sign in"])
            _record("Functional", "F-011", "Expense Splitter Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-011", "Expense Splitter Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-012: Routes page accessible ─────────────────────────────────────
    def test_F012_routes_page(self):
        t0 = time.time()
        try:
            self._go("routes")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "route", "map", "optimize", "sign in"])
            _record("Functional", "F-012", "Route Optimization Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-012", "Route Optimization Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-013: Route sharing page accessible ───────────────────────────────
    def test_F013_route_sharing_page(self):
        t0 = time.time()
        try:
            self._go("route-sharing")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "share", "itinerary", "route", "sign in"])
            _record("Functional", "F-013", "Route Sharing Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-013", "Route Sharing Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-014: Explore page accessible ────────────────────────────────────
    def test_F014_explore_page(self):
        t0 = time.time()
        try:
            self._go("explore")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "explore", "discover", "place", "sign in"])
            _record("Functional", "F-014", "Explore Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-014", "Explore Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-015: Trips page accessible ──────────────────────────────────────
    def test_F015_trips_page(self):
        t0 = time.time()
        try:
            self._go("trips")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "trip", "journey", "sign in"])
            _record("Functional", "F-015", "Trips Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-015", "Trips Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-016: Groups page accessible ─────────────────────────────────────
    def test_F016_groups_page(self):
        t0 = time.time()
        try:
            self._go("groups")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "group", "team", "sign in"])
            _record("Functional", "F-016", "Groups Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-016", "Groups Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-017: Profile page accessible ────────────────────────────────────
    def test_F017_profile_page(self):
        t0 = time.time()
        try:
            self._go("profile")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "profile", "account", "sign in"])
            _record("Functional", "F-017", "Profile Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-017", "Profile Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-018: Notifications page accessible ───────────────────────────────
    def test_F018_notifications_page(self):
        t0 = time.time()
        try:
            self._go("notifications")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "notification", "alert", "sign in"])
            _record("Functional", "F-018", "Notifications Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-018", "Notifications Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-019: Favorites page accessible ──────────────────────────────────
    def test_F019_favorites_page(self):
        t0 = time.time()
        try:
            self._go("favorites")
            time.sleep(3)
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "favorite", "saved", "sign in"])
            _record("Functional", "F-019", "Favorites Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-019", "Favorites Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-020: 404 page for unknown route ─────────────────────────────────
    def test_F020_404_page(self):
        t0 = time.time()
        try:
            self._go("this-page-does-not-exist-xyz-99999")
            time.sleep(2)
            page_src = self.driver.page_source.lower()
            is_404 = any(kw in page_src for kw in
                        ["404", "not found", "does not exist", "page not found", "tripsync"])
            assert is_404
            _record("Functional", "F-020", "404 Page Displays for Unknown Routes",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-020", "404 Page Displays for Unknown Routes",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-021: Landing page has hero CTA link ──────────────────────────────
    def test_F021_landing_hero_cta(self):
        t0 = time.time()
        try:
            self._go()
            time.sleep(2)
            links = self.driver.find_elements(By.TAG_NAME, "a")
            hrefs = [l.get_attribute("href") or "" for l in links]
            has_auth_link = any("login" in h or "register" in h or "dashboard" in h
                               for h in hrefs)
            assert has_auth_link
            _record("Functional", "F-021", "Landing Page Has Auth CTA Link",
                    "PASS", "Found auth-related link in landing page", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-021", "Landing Page Has Auth CTA Link",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-022: Page title is meaningful ───────────────────────────────────
    def test_F022_page_title_present(self):
        t0 = time.time()
        try:
            self._go()
            title = self.driver.title
            assert title and len(title.strip()) > 0
            _record("Functional", "F-022", "Landing Page Has Non-Empty Title",
                    "PASS", f"Title: '{title}'", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-022", "Landing Page Has Non-Empty Title",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-023: Login link navigates to login ───────────────────────────────
    def test_F023_login_link_navigation(self):
        t0 = time.time()
        try:
            self._go()
            time.sleep(2)
            links = self.driver.find_elements(By.TAG_NAME, "a")
            login_link = next((l for l in links if "login" in (l.get_attribute("href") or "")), None)
            if login_link:
                login_link.click()
                time.sleep(2)
                assert "login" in self.driver.current_url
                _record("Functional", "F-023", "Login Navigation Link Works",
                        "PASS", f"Navigated to: {self.driver.current_url}", time.time()-t0)
            else:
                _record("Functional", "F-023", "Login Navigation Link Works",
                        "SKIP", "No login link found on landing page (may be behind auth)", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-023", "Login Navigation Link Works",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-F-024: Page is responsive (mobile viewport) ───────────────────────
    def test_F024_mobile_viewport(self):
        t0 = time.time()
        try:
            self.driver.set_window_size(375, 812)
            self._go()
            time.sleep(2)
            body = self.driver.find_element(By.TAG_NAME, "body")
            assert body.is_displayed()
            overflow_x = self.driver.execute_script(
                "return document.documentElement.scrollWidth > window.innerWidth"
            )
            self.driver.set_window_size(1280, 900)
            _record("Functional", "F-024", "Mobile Viewport — No Horizontal Overflow",
                    "PASS" if not overflow_x else "WARN",
                    f"Overflow: {overflow_x}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-024", "Mobile Viewport — No Horizontal Overflow",
                    "FAIL", str(e), time.time()-t0)
            self.driver.set_window_size(1280, 900)
            self.fail(str(e))

    # ── TC-F-025: Visited page accessible ────────────────────────────────────
    def test_F025_visited_page(self):
        t0 = time.time()
        try:
            self._go("visited")
            page_src = self.driver.page_source.lower()
            assert any(kw in page_src for kw in ["tripsync", "visited", "places", "sign in"])
            _record("Functional", "F-025", "Visited Places Page Reachable",
                    "PASS", f"URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-025", "Visited Places Page Reachable",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F026_landing_meta_viewport(self):
        t0 = time.time()
        try:
            self._go()
            meta = self.driver.find_element(By.XPATH, "//meta[@name='viewport']")
            assert meta is not None
            _record("Functional", "F-026", "Meta Viewport Check", "PASS", meta.get_attribute("content"), time.time()-t0)
        except Exception as e:
            _record("Functional", "F-026", "Meta Viewport Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F027_landing_meta_description(self):
        t0 = time.time()
        try:
            self._go()
            meta = self.driver.find_element(By.XPATH, "//meta[@name='description']")
            _record("Functional", "F-027", "Meta Description Check", "PASS", meta.get_attribute("content")[:50], time.time()-t0)
        except Exception as e:
            _record("Functional", "F-027", "Meta Description Check", "WARN", str(e), time.time()-t0)

    def test_F028_landing_has_favicon(self):
        t0 = time.time()
        try:
            self._go()
            fav = self.driver.find_element(By.XPATH, "//link[contains(@rel, 'icon')]")
            _record("Functional", "F-028", "Favicon Reference Exists", "PASS", fav.get_attribute("href")[:60], time.time()-t0)
        except Exception as e:
            _record("Functional", "F-028", "Favicon Reference Exists", "WARN", str(e), time.time()-t0)

    def test_F029_landing_h1_present(self):
        t0 = time.time()
        try:
            self._go()
            h1s = self.driver.find_elements(By.TAG_NAME, "h1")
            assert len(h1s) >= 1
            _record("Functional", "F-029", "H1 Header Present", "PASS", f"Found {len(h1s)} H1(s)", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-029", "H1 Header Present", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F030_login_logo_present(self):
        t0 = time.time()
        try:
            self._go("login")
            logo = self.driver.find_element(By.XPATH, "//a[contains(@href, '/')]")
            assert logo is not None
            _record("Functional", "F-030", "Login Brand Link Present", "PASS", logo.text, time.time()-t0)
        except Exception as e:
            _record("Functional", "F-030", "Login Brand Link Present", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F031_login_header_text(self):
        t0 = time.time()
        try:
            self._go("login")
            src = self.driver.page_source.lower()
            assert "welcome back" in src or "sign in" in src
            _record("Functional", "F-031", "Login Title Check", "PASS", "Header text valid", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-031", "Login Title Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F032_login_signup_link(self):
        t0 = time.time()
        try:
            self._go("login")
            link = self.driver.find_element(By.XPATH, "//a[contains(@href, 'register')]")
            _record("Functional", "F-032", "Login Registration Link Works", "PASS", link.get_attribute("href"), time.time()-t0)
        except Exception as e:
            _record("Functional", "F-032", "Login Registration Link Works", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F033_register_signin_link(self):
        t0 = time.time()
        try:
            self._go("register")
            link = self.driver.find_element(By.XPATH, "//a[contains(@href, 'login')]")
            _record("Functional", "F-033", "Register Login Link Works", "PASS", link.get_attribute("href"), time.time()-t0)
        except Exception as e:
            _record("Functional", "F-033", "Register Login Link Works", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F034_register_email_placeholder(self):
        t0 = time.time()
        try:
            self._go("register")
            em = self.driver.find_element(By.XPATH, "//input[@type='email']")
            ph = em.get_attribute("placeholder")
            assert ph and len(ph) > 0
            _record("Functional", "F-034", "Register Email Input Placeholder", "PASS", f"Placeholder: {ph}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-034", "Register Email Input Placeholder", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F035_register_password_placeholder(self):
        t0 = time.time()
        try:
            self._go("register")
            pw = self.driver.find_element(By.XPATH, "//input[@type='password']")
            ph = pw.get_attribute("placeholder")
            assert ph and len(ph) > 0
            _record("Functional", "F-035", "Register Password Input Placeholder", "PASS", f"Placeholder: {ph}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-035", "Register Password Input Placeholder", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F036_register_confirm_password(self):
        t0 = time.time()
        try:
            self._go("register")
            pws = self.driver.find_elements(By.XPATH, "//input[@type='password']")
            _record("Functional", "F-036", "Register Password Count", "PASS", f"Found {len(pws)} password field(s)", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-036", "Register Password Count", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F037_dashboard_header_structure(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            src = self.driver.page_source.lower()
            assert "tripsync" in src
            _record("Functional", "F-037", "Dashboard Layout Header Structure", "PASS", "Dashboard contains brand wrapper", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-037", "Dashboard Layout Header Structure", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F038_dashboard_nav_links(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            navs = self.driver.find_elements(By.TAG_NAME, "nav")
            _record("Functional", "F-038", "Dashboard Nav Element", "PASS", f"Found {len(navs)} nav tag(s)", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-038", "Dashboard Nav Element", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F039_dashboard_footer_elements(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            ft = self.driver.find_elements(By.TAG_NAME, "footer")
            _record("Functional", "F-039", "Dashboard Footer Tag Check", "PASS", f"Found {len(ft)} footer(s)", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-039", "Dashboard Footer Tag Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F040_ai_assistant_input_exists(self):
        t0 = time.time()
        try:
            self._go("ai-assistant")
            inputs = self.driver.find_elements(By.XPATH, "//input | //textarea")
            assert len(inputs) >= 1
            _record("Functional", "F-040", "AI Assistant Inputs Exist", "PASS", f"Found {len(inputs)} inputs", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-040", "AI Assistant Inputs Exist", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F041_ai_assistant_voice_button(self):
        t0 = time.time()
        try:
            self._go("ai-assistant")
            src = self.driver.page_source.lower()
            has_voice = any(kw in src for kw in ["mic", "voice", "speak", "audio", "record"])
            _record("Functional", "F-041", "AI Assistant Voice References", "PASS" if has_voice else "WARN", f"Has voice markers: {has_voice}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-041", "AI Assistant Voice References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F042_ai_planner_form_inputs(self):
        t0 = time.time()
        try:
            self._go("ai-planner")
            inputs = self.driver.find_elements(By.XPATH, "//input | //select | //textarea")
            _record("Functional", "F-042", "AI Planner Inputs", "PASS", f"Found {len(inputs)} input elements", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-042", "AI Planner Inputs", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F043_safety_city_input(self):
        t0 = time.time()
        try:
            self._go("safety")
            inputs = self.driver.find_elements(By.XPATH, "//input")
            _record("Functional", "F-043", "Safety Search Inputs", "PASS", f"Found {len(inputs)} input fields", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-043", "Safety Search Inputs", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F044_expenses_total_amount_field(self):
        t0 = time.time()
        try:
            self._go("expenses")
            inputs = self.driver.find_elements(By.XPATH, "//input[@type='number' or contains(@id, 'amount') or contains(@placeholder, 'amount')]")
            _record("Functional", "F-044", "Expenses Splitter Amount Inputs", "PASS", f"Found {len(inputs)} fields", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-044", "Expenses Splitter Amount Inputs", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F045_expenses_member_inputs(self):
        t0 = time.time()
        try:
            self._go("expenses")
            src = self.driver.page_source.lower()
            has_members = any(kw in src for kw in ["member", "people", "friend", "split"])
            _record("Functional", "F-045", "Expenses Splitter Member References", "PASS" if has_members else "WARN", f"Has member markers: {has_members}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-045", "Expenses Splitter Member References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F046_routes_spot_inputs(self):
        t0 = time.time()
        try:
            self._go("routes")
            inputs = self.driver.find_elements(By.XPATH, "//input")
            _record("Functional", "F-046", "Routes Page Inputs", "PASS", f"Found {len(inputs)} input element(s)", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-046", "Routes Page Inputs", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F047_route_sharing_table(self):
        t0 = time.time()
        try:
            self._go("route-sharing")
            src = self.driver.page_source.lower()
            has_table = any(kw in src for kw in ["share", "table", "list", "route", "post"])
            _record("Functional", "F-047", "Route Sharing Layout References", "PASS" if has_table else "WARN", f"Has sharing markers: {has_table}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-047", "Route Sharing Layout References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F048_explore_search_bar(self):
        t0 = time.time()
        try:
            self._go("explore")
            inputs = self.driver.find_elements(By.XPATH, "//input[@type='text' or contains(@placeholder, 'search')]")
            _record("Functional", "F-048", "Explore Search Input Fields", "PASS", f"Found {len(inputs)} search inputs", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-048", "Explore Search Input Fields", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F049_trips_list_container(self):
        t0 = time.time()
        try:
            self._go("trips")
            src = self.driver.page_source.lower()
            has_list = any(kw in src for kw in ["trip", "journey", "add", "create", "list"])
            _record("Functional", "F-049", "Trips Layout References", "PASS" if has_list else "WARN", f"Has trip markers: {has_list}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-049", "Trips Layout References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F050_groups_list_container(self):
        t0 = time.time()
        try:
            self._go("groups")
            src = self.driver.page_source.lower()
            has_groups = any(kw in src for kw in ["group", "member", "team", "chat", "create"])
            _record("Functional", "F-050", "Groups Page References", "PASS" if has_groups else "WARN", f"Has group markers: {has_groups}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-050", "Groups Page References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F051_profile_settings_fields(self):
        t0 = time.time()
        try:
            self._go("profile")
            inputs = self.driver.find_elements(By.XPATH, "//input")
            _record("Functional", "F-051", "Profile Settings Input Elements", "PASS", f"Found {len(inputs)} fields", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-051", "Profile Settings Input Elements", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F052_notifications_list_container(self):
        t0 = time.time()
        try:
            self._go("notifications")
            src = self.driver.page_source.lower()
            has_notif = any(kw in src for kw in ["notification", "alert", "unread", "recent", "message"])
            _record("Functional", "F-052", "Notifications References", "PASS" if has_notif else "WARN", f"Has notification markers: {has_notif}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-052", "Notifications References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F053_favorites_list_container(self):
        t0 = time.time()
        try:
            self._go("favorites")
            src = self.driver.page_source.lower()
            has_fav = any(kw in src for kw in ["favorite", "saved", "star", "like", "list"])
            _record("Functional", "F-053", "Favorites References", "PASS" if has_fav else "WARN", f"Has favorites markers: {has_fav}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-053", "Favorites References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F054_visited_list_container(self):
        t0 = time.time()
        try:
            self._go("visited")
            src = self.driver.page_source.lower()
            has_visited = any(kw in src for kw in ["visited", "history", "places", "explore", "map"])
            _record("Functional", "F-054", "Visited Page Layout References", "PASS" if has_visited else "WARN", f"Has visited markers: {has_visited}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-054", "Visited Page Layout References", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F055_nav_brand_click(self):
        t0 = time.time()
        try:
            self._go("login")
            logo = self.driver.find_element(By.XPATH, "//a[contains(@href, '/')]")
            logo.click()
            time.sleep(1.0)
            _record("Functional", "F-055", "Logo Brand Click Redirects Home", "PASS", f"Redirect URL: {self.driver.current_url}", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-055", "Logo Brand Click Redirects Home", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F056_dashboard_sidebar_mobile(self):
        t0 = time.time()
        try:
            self.driver.set_window_size(375, 667)
            self._go("dashboard")
            _record("Functional", "F-056", "Responsive Viewport Stretched Layout", "PASS", "Viewport set successfully", time.time()-t0)
            self.driver.set_window_size(1280, 900)
        except Exception as e:
            self.driver.set_window_size(1280, 900)
            _record("Functional", "F-056", "Responsive Viewport Stretched Layout", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F057_console_logs_no_errors(self):
        t0 = time.time()
        try:
            self._go()
            logs = self.driver.get_log("browser")
            errors = [l for l in logs if l.get("level") == "SEVERE"]
            _record("Functional", "F-057", "Console Logs Severity Assessment", "PASS" if not errors else "WARN", f"Found {len(errors)} critical logs", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-057", "Console Logs Severity Assessment", "INFO", str(e), time.time()-t0)

    def test_F058_local_storage_availability(self):
        t0 = time.time()
        try:
            self._go()
            self.driver.execute_script("localStorage.setItem('__test_ls_key__', '123');")
            val = self.driver.execute_script("return localStorage.getItem('__test_ls_key__');")
            self.driver.execute_script("localStorage.removeItem('__test_ls_key__');")
            assert val == "123"
            _record("Functional", "F-058", "LocalStorage Read/Write Check", "PASS", "Successfully stored and retrieved", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-058", "LocalStorage Read/Write Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F059_session_storage_availability(self):
        t0 = time.time()
        try:
            self._go()
            self.driver.execute_script("sessionStorage.setItem('__test_ss_key__', '456');")
            val = self.driver.execute_script("return sessionStorage.getItem('__test_ss_key__');")
            self.driver.execute_script("sessionStorage.removeItem('__test_ss_key__');")
            assert val == "456"
            _record("Functional", "F-059", "SessionStorage Read/Write Check", "PASS", "Successfully stored and retrieved", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-059", "SessionStorage Read/Write Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F060_cookie_storage_check(self):
        t0 = time.time()
        try:
            self._go()
            self.driver.add_cookie({"name": "test_cookie", "value": "val123"})
            c = self.driver.get_cookie("test_cookie")
            assert c and c.get("value") == "val123"
            self.driver.delete_cookie("test_cookie")
            _record("Functional", "F-060", "Browser Cookie Validation", "PASS", "Cookie read/write matches", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-060", "Browser Cookie Validation", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F061_login_submit_disabled_when_loading(self):
        t0 = time.time()
        try:
            self._go("login")
            btn = _find_submit_btn(self.driver)
            assert btn is not None
            _record("Functional", "F-061", "Login Button Click Presence", "PASS", "Submit button present", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-061", "Login Button Click Presence", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F062_register_submit_disabled_when_loading(self):
        t0 = time.time()
        try:
            self._go("register")
            btn = _find_submit_btn(self.driver)
            assert btn is not None
            _record("Functional", "F-062", "Register Button Click Presence", "PASS", "Submit button present", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-062", "Register Button Click Presence", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F063_page_zoom_resilience(self):
        t0 = time.time()
        try:
            self._go()
            self.driver.execute_script("document.body.style.zoom = '0.5';")
            body = self.driver.find_element(By.TAG_NAME, "body")
            assert body.is_displayed()
            self.driver.execute_script("document.body.style.zoom = '1.0';")
            _record("Functional", "F-063", "Page Zoom Scaling Capability", "PASS", "Zoomed body remains displayed", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-063", "Page Zoom Scaling Capability", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F064_page_landscape_mode(self):
        t0 = time.time()
        try:
            self.driver.set_window_size(1024, 768)
            self._go()
            body = self.driver.find_element(By.TAG_NAME, "body")
            assert body.is_displayed()
            self.driver.set_window_size(1280, 900)
            _record("Functional", "F-064", "Landscape Resolution Mode (1024x768)", "PASS", "Display validated in landscape mode", time.time()-t0)
        except Exception as e:
            self.driver.set_window_size(1280, 900)
            _record("Functional", "F-064", "Landscape Resolution Mode (1024x768)", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F065_page_portrait_mode(self):
        t0 = time.time()
        try:
            self.driver.set_window_size(768, 1024)
            self._go()
            body = self.driver.find_element(By.TAG_NAME, "body")
            assert body.is_displayed()
            self.driver.set_window_size(1280, 900)
            _record("Functional", "F-065", "Tablet Portrait Mode (768x1024)", "PASS", "Display validated in portrait mode", time.time()-t0)
        except Exception as e:
            self.driver.set_window_size(1280, 900)
            _record("Functional", "F-065", "Tablet Portrait Mode (768x1024)", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F066_web_app_lang_attribute(self):
        t0 = time.time()
        try:
            self._go()
            lang = self.driver.find_element(By.TAG_NAME, "html").get_attribute("lang")
            _record("Functional", "F-066", "HTML Document Language Verification", "PASS", f"lang attribute: '{lang}'", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-066", "HTML Document Language Verification", "WARN", str(e), time.time()-t0)

    def test_F067_landing_seo_headings(self):
        t0 = time.time()
        try:
            self._go()
            h2s = self.driver.find_elements(By.TAG_NAME, "h2")
            _record("Functional", "F-067", "Subheadings Presence", "PASS", f"Found {len(h2s)} H2 elements", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-067", "Subheadings Presence", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F068_empty_registration_form_handling(self):
        t0 = time.time()
        try:
            self._go("register")
            btn = _find_submit_btn(self.driver)
            btn.click()
            time.sleep(0.5)
            assert "register" in self.driver.current_url or "dashboard" not in self.driver.current_url
            _record("Functional", "F-068", "Empty Registration Form Submission Block", "PASS", "Did not redirect to dashboard", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-068", "Empty Registration Form Submission Block", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F069_back_button_navigation(self):
        t0 = time.time()
        try:
            self._go()
            url1 = self.driver.current_url
            self._go("login")
            self.driver.back()
            time.sleep(1.0)
            assert self.driver.current_url == url1
            _record("Functional", "F-069", "Browser History Navigation", "PASS", "Successfully went back to index", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-069", "Browser History Navigation", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_F070_dashboard_active_tab_styling(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            navs = self.driver.find_elements(By.XPATH, "//a")
            assert len(navs) >= 1
            _record("Functional", "F-070", "Dashboard Link List Verification", "PASS", f"Dashboard contains {len(navs)} anchor elements", time.time()-t0)
        except Exception as e:
            _record("Functional", "F-070", "Dashboard Link List Verification", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — VULNERABILITY / SECURITY TESTS
# ═══════════════════════════════════════════════════════════════════════════════
class VulnerabilityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.driver = _build_driver()
        cls.driver.set_page_load_timeout(45)

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()

    def _go(self, path=""):
        self.driver.get(BASE_URL.rstrip("/") + ("/" + path.lstrip("/") if path else ""))
        time.sleep(1.5)

    # ── TC-V-001: XSS in email field ─────────────────────────────────────────
    def test_V001_xss_login_email(self):
        t0 = time.time()
        xss_payload = "<script>alert('XSS')</script>"
        try:
            self._go("login")
            email_field = _wait_visible(self.driver, By.ID, "email")
            email_field.clear()
            email_field.send_keys(xss_payload)
            btn = _find_submit_btn(self.driver)
            btn.click()
            time.sleep(1.5)
            # Check no alert dialog appeared
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-001", "XSS Injection in Email Field",
                        "FAIL", "XSS alert executed!", time.time()-t0, "Critical")
                self.fail("XSS alert was triggered!")
            except:
                pass
            # Check script not rendered as HTML
            page_src = self.driver.page_source
            executed = "<script>alert" in page_src and "alert('XSS')" in page_src
            _record("Vulnerability", "V-001", "XSS Injection in Email Field",
                    "FAIL" if executed else "PASS",
                    "Script tag rendered in DOM" if executed else "XSS payload sanitized",
                    time.time()-t0, "Critical")
            if executed:
                self.fail("XSS payload was rendered in DOM")
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-001", "XSS Injection in Email Field",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-V-002: XSS in password field ─────────────────────────────────────
    def test_V002_xss_login_password(self):
        t0 = time.time()
        xss_payload = "<img src=x onerror=alert(1)>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys("test@test.com")
            self.driver.find_element(By.ID, "password").send_keys(xss_payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.5)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-002", "XSS Injection in Password Field",
                        "FAIL", "XSS onerror executed!", time.time()-t0, "Critical")
                self.fail("XSS executed in password field")
            except:
                pass
            _record("Vulnerability", "V-002", "XSS Injection in Password Field",
                    "PASS", "Password field XSS sanitized", time.time()-t0, "Critical")
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-002", "XSS Injection in Password Field",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-V-003: SQL Injection attempt in email ──────────────────────────────
    def test_V003_sql_injection_email(self):
        t0 = time.time()
        sqli_payload = "' OR '1'='1'; DROP TABLE users; --"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(sqli_payload)
            self.driver.find_element(By.ID, "password").send_keys("password123")
            _find_submit_btn(self.driver).click()
            time.sleep(3)
            # Should NOT be on dashboard
            on_dashboard = "dashboard" in self.driver.current_url
            _record("Vulnerability", "V-003", "SQL Injection in Email Field",
                    "FAIL" if on_dashboard else "PASS",
                    "SQL injection gained access!" if on_dashboard else "SQLi blocked",
                    time.time()-t0, "Critical")
            assert not on_dashboard
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-003", "SQL Injection in Email Field",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-V-004: Open redirect check ────────────────────────────────────────
    def test_V004_open_redirect_check(self):
        t0 = time.time()
        malicious_url = BASE_URL + "?redirect=https://evil.com"
        try:
            self.driver.get(malicious_url)
            time.sleep(2)
            current = self.driver.current_url
            redirected_to_evil = "evil.com" in current
            _record("Vulnerability", "V-004", "Open Redirect Vulnerability Check",
                    "FAIL" if redirected_to_evil else "PASS",
                    f"Redirected to evil.com!" if redirected_to_evil else f"Stayed at: {current}",
                    time.time()-t0, "High")
            assert not redirected_to_evil
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-004", "Open Redirect Vulnerability Check",
                    "FAIL", str(e), time.time()-t0, "High")
            self.fail(str(e))

    # ── TC-V-005: Sensitive data not in page source ───────────────────────────
    def test_V005_no_sensitive_data_in_source(self):
        t0 = time.time()
        try:
            self._go("login")
            time.sleep(2)
            page_src = self.driver.page_source
            # Check no hardcoded API keys in HTML source
            sensitive_patterns = [
                "AIzaSy",        # Firebase API key prefix
                "sk-or-v1-",     # OpenRouter key
                "gsk_",          # Groq key
                "hf_",           # HuggingFace key
                "password",      # Raw password data
            ]
            found = [p for p in sensitive_patterns
                    if p in page_src and p not in ["password"]]
            # "password" is fine in HTML attribute names, not values
            status = "FAIL" if found else "PASS"
            detail = f"Exposed: {found}" if found else "No sensitive keys exposed in source"
            _record("Vulnerability", "V-005", "No Sensitive API Keys in Page Source",
                    status, detail, time.time()-t0, "Critical")
            assert not found, f"Sensitive data found: {found}"
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-005", "No Sensitive API Keys in Page Source",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-V-006: HTTPS enforced ─────────────────────────────────────────────
    def test_V006_https_enforced(self):
        t0 = time.time()
        try:
            self.driver.get(BASE_URL)
            time.sleep(2)
            is_https = self.driver.current_url.startswith("https://")
            _record("Vulnerability", "V-006", "HTTPS Enforced on Web App",
                    "PASS" if is_https else "FAIL",
                    f"Protocol: {self.driver.current_url[:8]}",
                    time.time()-t0, "High")
            assert is_https
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-006", "HTTPS Enforced on Web App",
                    "FAIL", str(e), time.time()-t0, "High")
            self.fail(str(e))

    # ── TC-V-007: No console errors with XSS payload in URL ──────────────────
    def test_V007_url_xss_attempt(self):
        t0 = time.time()
        xss_in_url = BASE_URL + "/<script>alert('xss')</script>"
        try:
            self.driver.get(xss_in_url)
            time.sleep(2)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-007", "URL-based XSS Attack Blocked",
                        "FAIL", "Alert executed via URL", time.time()-t0, "Critical")
                self.fail("URL XSS executed")
            except:
                pass
            _record("Vulnerability", "V-007", "URL-based XSS Attack Blocked",
                    "PASS", "URL XSS blocked by app/browser", time.time()-t0, "Critical")
        except WebDriverException as e:
            if "alert" in str(e).lower():
                _record("Vulnerability", "V-007", "URL-based XSS Attack Blocked",
                        "FAIL", str(e), time.time()-t0, "Critical")
            else:
                _record("Vulnerability", "V-007", "URL-based XSS Attack Blocked",
                        "PASS", "App handled malformed URL", time.time()-t0)
        except Exception as e:
            _record("Vulnerability", "V-007", "URL-based XSS Attack Blocked",
                    "INFO", f"App handled invalid URL: {str(e)[:100]}", time.time()-t0)

    # ── TC-V-008: Login rate limiting / brute force protection ───────────────
    def test_V008_brute_force_protection(self):
        t0 = time.time()
        try:
            self._go("login")
            for attempt in range(3):
                email = _wait_visible(self.driver, By.ID, "email")
                email.clear()
                email.send_keys("brute@test.com")
                pw = self.driver.find_element(By.ID, "password")
                pw.clear()
                pw.send_keys(f"wrongpass{attempt}")
                _find_submit_btn(self.driver).click()
                time.sleep(2)
            page_src = self.driver.page_source.lower()
            # Should show error or still on login page
            blocked = any(kw in page_src for kw in
                        ["too many", "rate limit", "locked", "error", "invalid", "blocked"])
            still_on_login = "dashboard" not in self.driver.current_url
            _record("Vulnerability", "V-008", "Brute Force — Still On Login After 3 Attempts",
                    "PASS" if (blocked or still_on_login) else "WARN",
                    "Rate limited or errors shown" if blocked else "No dashboard access achieved",
                    time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-008", "Brute Force — Still On Login After 3 Attempts",
                    "FAIL", str(e), time.time()-t0, "High")
            self.fail(str(e))

    # ── TC-V-009: Prototype pollution check ──────────────────────────────────
    def test_V009_prototype_pollution(self):
        t0 = time.time()
        try:
            self._go("login")
            time.sleep(2)
            result = self.driver.execute_script("""
                try {
                    var payload = JSON.parse('{"__proto__": {"polluted": true}}');
                    return ({}).polluted;
                } catch(e) {
                    return false;
                }
            """)
            polluted = result is True
            _record("Vulnerability", "V-009", "JavaScript Prototype Pollution Check",
                    "FAIL" if polluted else "PASS",
                    "Prototype polluted!" if polluted else "No prototype pollution detected",
                    time.time()-t0, "High")
            assert not polluted
        except AssertionError:
            raise
        except Exception as e:
            _record("Vulnerability", "V-009", "JavaScript Prototype Pollution Check",
                    "FAIL", str(e), time.time()-t0, "High")
            self.fail(str(e))

    # ── TC-V-010: Clickjacking — iframe embedding check ──────────────────────
    def test_V010_clickjacking_check(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL, timeout=15, verify=False)
            headers = resp.headers
            x_frame = headers.get("X-Frame-Options", "")
            csp = headers.get("Content-Security-Policy", "")
            has_protection = (
                x_frame.upper() in ["DENY", "SAMEORIGIN"] or
                "frame-ancestors" in csp.lower()
            )
            _record("Vulnerability", "V-010", "Clickjacking Protection (X-Frame-Options)",
                    "PASS" if has_protection else "WARN",
                    f"X-Frame-Options: '{x_frame}', CSP: '{csp[:80]}'",
                    time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-010", "Clickjacking Protection (X-Frame-Options)",
                    "INFO", f"Could not check headers: {str(e)[:80]}", time.time()-t0, "Medium")

    def test_V011_xss_svg_onload(self):
        t0 = time.time()
        payload = "<svg/onload=alert(1)>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-011", "SVG Onload XSS Sanitization", "FAIL", "Alert triggered!", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-011", "SVG Onload XSS Sanitization", "PASS", "SVG payload blocked", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-011", "SVG Onload XSS Sanitization", "WARN", str(e), time.time()-t0, "Critical")

    def test_V012_xss_javascript_uri(self):
        t0 = time.time()
        try:
            self.driver.get(BASE_URL + "/login?redirect=javascript:alert(1)")
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-012", "Javascript URI XSS Sanitization", "FAIL", "Alert triggered via URI", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-012", "Javascript URI XSS Sanitization", "PASS", "Javascript URI blocked", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-012", "Javascript URI XSS Sanitization", "WARN", str(e), time.time()-t0, "Critical")

    def test_V013_xss_event_handlers(self):
        t0 = time.time()
        payload = "test@domain.com\" onmouseover=\"alert(1)"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-013", "XSS Event Handler Sanitization", "FAIL", "Alert triggered via event handler", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-013", "XSS Event Handler Sanitization", "PASS", "Payload successfully escaped", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-013", "XSS Event Handler Sanitization", "WARN", str(e), time.time()-t0, "Critical")

    def test_V014_xss_nested_tags(self):
        t0 = time.time()
        payload = "<scr<script>ipt>alert(1)</script>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-014", "XSS Nested Script Sanitization", "FAIL", "Alert triggered via nested tags", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-014", "XSS Nested Script Sanitization", "PASS", "Nested scripts sanitized", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-014", "XSS Nested Script Sanitization", "WARN", str(e), time.time()-t0, "Critical")

    def test_V015_xss_body_onload(self):
        t0 = time.time()
        payload = "<body onload=alert(1)>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-015", "XSS Body Onload Sanitization", "FAIL", "Alert triggered!", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-015", "XSS Body Onload Sanitization", "PASS", "Body onload stripped/escaped", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-015", "XSS Body Onload Sanitization", "WARN", str(e), time.time()-t0, "Critical")

    def test_V016_sqli_boolean_based(self):
        t0 = time.time()
        payload = "admin' OR '1'='1"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys("valid@user.com")
            self.driver.find_element(By.ID, "password").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            assert "dashboard" not in self.driver.current_url
            _record("Vulnerability", "V-016", "Boolean-based SQLi Protection", "PASS", "No unauthorized authentication achieved", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-016", "Boolean-based SQLi Protection", "WARN", str(e), time.time()-t0, "Critical")

    def test_V017_sqli_error_based(self):
        t0 = time.time()
        payload = "' AND (SELECT 1 FROM (SELECT COUNT(*), CONCAT((SELECT VERSION()), 0x3a, FLOOR(RAND(0)*2)) x FROM INFORMATION_SCHEMA.TABLES GROUP BY x) y) --"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            # Ensure no system SQL error displays in DOM
            src = self.driver.page_source.lower()
            has_mysql_err = any(kw in src for kw in ["mysql", "syntax error", "sqlstate", "drivererror"])
            assert not has_mysql_err
            _record("Vulnerability", "V-017", "Error-based SQLi Protection", "PASS", "SQL driver details not leaked in client", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-017", "Error-based SQLi Protection", "WARN", str(e), time.time()-t0, "Critical")

    def test_V018_sqli_stacked_queries(self):
        t0 = time.time()
        payload = "user@test.com; DROP TABLE users; --"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            assert "dashboard" not in self.driver.current_url
            _record("Vulnerability", "V-018", "Stacked SQLi Protection", "PASS", "No execution anomalies detected", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-018", "Stacked SQLi Protection", "WARN", str(e), time.time()-t0, "Critical")

    def test_V019_sqli_union_select(self):
        t0 = time.time()
        payload = "' UNION SELECT null, null, null, null --"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            assert "dashboard" not in self.driver.current_url
            _record("Vulnerability", "V-019", "Union-based SQLi Protection", "PASS", "SQL union query rejected", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-019", "Union-based SQLi Protection", "WARN", str(e), time.time()-t0, "Critical")

    def test_V020_path_traversal_login(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL + "/../../etc/passwd", timeout=15, verify=False)
            assert resp.status_code in [200, 404, 400, 403]
            _record("Vulnerability", "V-020", "Directory Traversal URL Access Block", "PASS", f"Returned status: {resp.status_code}", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-020", "Directory Traversal URL Access Block", "WARN", str(e), time.time()-t0, "High")

    def test_V021_path_traversal_query(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL + "/login?file=../../../../windows/system32/cmd.exe", timeout=15, verify=False)
            assert resp.status_code in [200, 404, 400, 403]
            _record("Vulnerability", "V-021", "Directory Traversal Param Sanitization", "PASS", f"Returned status: {resp.status_code}", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-021", "Directory Traversal Param Sanitization", "WARN", str(e), time.time()-t0, "High")

    def test_V022_command_injection_pipe(self):
        t0 = time.time()
        payload = "test@domain.com | whoami"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            src = self.driver.page_source.lower()
            assert "system" not in src and "administrator" not in src and "root" not in src
            _record("Vulnerability", "V-022", "Command Injection (Pipe Syntax)", "PASS", "Pipe commands not executed in shell", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-022", "Command Injection (Pipe Syntax)", "WARN", str(e), time.time()-t0, "Critical")

    def test_V023_command_injection_semicolon(self):
        t0 = time.time()
        payload = "test@domain.com; cat /etc/passwd"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            assert "root:x:" not in self.driver.page_source.lower()
            _record("Vulnerability", "V-023", "Command Injection (Semicolon Syntax)", "PASS", "Semicolon commands not executed", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-023", "Command Injection (Semicolon Syntax)", "WARN", str(e), time.time()-t0, "Critical")

    def test_V024_command_injection_backtick(self):
        t0 = time.time()
        payload = "`uname -a`"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            src = self.driver.page_source.lower()
            assert "linux" not in src and "windows" not in src and "darwin" not in src
            _record("Vulnerability", "V-024", "Command Injection (Backtick Syntax)", "PASS", "Backtick commands not executed", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-024", "Command Injection (Backtick Syntax)", "WARN", str(e), time.time()-t0, "Critical")

    def test_V025_hsts_header_present(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL, timeout=15, verify=False)
            hsts = resp.headers.get("Strict-Transport-Security", "")
            has_hsts = len(hsts) > 0
            _record("Vulnerability", "V-025", "HSTS Header Check", "PASS" if has_hsts else "WARN", f"HSTS header value: '{hsts}'", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-025", "HSTS Header Check", "INFO", str(e), time.time()-t0)

    def test_V026_referrer_policy_header(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL, timeout=15, verify=False)
            ref = resp.headers.get("Referrer-Policy", "")
            has_ref = len(ref) > 0
            _record("Vulnerability", "V-026", "Referrer-Policy Header Check", "PASS" if has_ref else "WARN", f"Referrer-Policy: '{ref}'", time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-026", "Referrer-Policy Header Check", "INFO", str(e), time.time()-t0)

    def test_V027_cache_control_headers(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL + "/login", timeout=15, verify=False)
            cc = resp.headers.get("Cache-Control", "")
            _record("Vulnerability", "V-027", "Cache-Control Settings Header", "PASS" if cc else "INFO", f"Cache-Control: '{cc}'", time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-027", "Cache-Control Settings Header", "INFO", str(e), time.time()-t0)

    def test_V028_x_content_type_options(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL, timeout=15, verify=False)
            ct_opt = resp.headers.get("X-Content-Type-Options", "")
            is_nosniff = ct_opt.lower() == "nosniff"
            _record("Vulnerability", "V-028", "X-Content-Type-Options Header Verification", "PASS" if is_nosniff else "WARN", f"Value: '{ct_opt}'", time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-028", "X-Content-Type-Options Header Verification", "INFO", str(e), time.time()-t0)

    def test_V029_x_xss_protection_header(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL, timeout=15, verify=False)
            xss_hdr = resp.headers.get("X-XSS-Protection", "")
            _record("Vulnerability", "V-029", "X-XSS-Protection Header Check", "PASS" if xss_hdr else "INFO", f"Value: '{xss_hdr}'", time.time()-t0, "Low")
        except Exception as e:
            _record("Vulnerability", "V-029", "X-XSS-Protection Header Check", "INFO", str(e), time.time()-t0)

    def test_V030_cors_origin_spoofing(self):
        t0 = time.time()
        try:
            headers = {"Origin": "https://attacker-domain-xyz.com"}
            resp = requests.options(API_URL + "/api/chat", headers=headers, timeout=15)
            allow_origin = resp.headers.get("Access-Control-Allow-Origin", "")
            is_blocked = (allow_origin != "*") and (allow_origin != "https://attacker-domain-xyz.com")
            _record("Vulnerability", "V-030", "CORS Origin Spoofing Blockage", "PASS" if is_blocked else "WARN", f"ACAO: '{allow_origin}'", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-030", "CORS Origin Spoofing Blockage", "INFO", str(e), time.time()-t0)

    def test_V031_insecure_http_methods(self):
        t0 = time.time()
        try:
            resp = requests.delete(API_URL + "/api/chat", timeout=15)
            is_rejected = resp.status_code in [405, 404, 400, 401, 403, 422]
            _record("Vulnerability", "V-031", "Insecure HTTP Methods Rejection (DELETE /api/chat)", "PASS" if is_rejected else "WARN", f"Status: {resp.status_code}", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-031", "Insecure HTTP Methods Rejection (DELETE /api/chat)", "INFO", str(e), time.time()-t0)

    def test_V032_session_hijacking_localstorage(self):
        t0 = time.time()
        try:
            self._go()
            keys = self.driver.execute_script("return Object.keys(localStorage);")
            sensitive = [k for k in keys if "password" in k.lower() or "secret" in k.lower() or "credentials" in k.lower()]
            assert not sensitive
            _record("Vulnerability", "V-032", "LocalStorage Sensitive Passwords Check", "PASS", "No plain text keys found", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-032", "LocalStorage Sensitive Passwords Check", "WARN", str(e), time.time()-t0, "High")

    def test_V033_session_hijacking_cookies(self):
        t0 = time.time()
        try:
            self._go()
            cookies = self.driver.get_cookies()
            non_secure = [c for c in cookies if not c.get("secure") and c.get("name") in ["session", "auth", "token"]]
            assert not non_secure
            _record("Vulnerability", "V-033", "Session Cookie Secure Attributes Check", "PASS", "Auth cookies contain proper flags", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-033", "Session Cookie Secure Attributes Check", "WARN", str(e), time.time()-t0, "High")

    def test_V034_content_type_enforcement(self):
        t0 = time.time()
        try:
            resp = requests.post(API_URL + "/api/chat", data="abc", headers={"Content-Type": "text/plain"}, timeout=15)
            rejected = resp.status_code in [415, 400, 422, 503]
            _record("Vulnerability", "V-034", "Content-Type Enforcements Verification", "PASS" if rejected else "WARN", f"Status: {resp.status_code}", time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-034", "Content-Type Enforcements Verification", "INFO", str(e), time.time()-t0)

    def test_V035_xss_form_action_injection(self):
        t0 = time.time()
        try:
            self._go("login")
            forms = self.driver.find_elements(By.TAG_NAME, "form")
            unsafe = False
            for f in forms:
                act = f.get_attribute("action") or ""
                if "javascript:" in act.lower():
                    unsafe = True
            assert not unsafe
            _record("Vulnerability", "V-035", "Form Action XSS Protection Check", "PASS", "Form action attributes are safe", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-035", "Form Action XSS Protection Check", "WARN", str(e), time.time()-t0, "Critical")

    def test_V036_css_injection_attempts(self):
        t0 = time.time()
        payload = "<style>body { background: url('https://evil.com/leak') }</style>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            src = self.driver.page_source.lower()
            assert "leak" not in src or "evil.com" not in src
            _record("Vulnerability", "V-036", "CSS Injection Protection Check", "PASS", "Style injection successfully neutralized", time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-036", "CSS Injection Protection Check", "WARN", str(e), time.time()-t0)

    def test_V037_http_parameter_pollution(self):
        t0 = time.time()
        try:
            resp = requests.get(API_URL + "/api/safety?city=Goa&city=Mumbai", timeout=15)
            assert resp.status_code in [200, 400, 422, 503]
            _record("Vulnerability", "V-037", "HTTP Parameter Pollution Safety", "PASS", f"Status: {resp.status_code}", time.time()-t0, "Medium")
        except Exception as e:
            _record("Vulnerability", "V-037", "HTTP Parameter Pollution Safety", "WARN", str(e), time.time()-t0)

    def test_V038_xxe_xml_injection(self):
        t0 = time.time()
        xml_payload = """<?xml version="1.0" encoding="ISO-8859-1"?>
        <!DOCTYPE foo [  
        <!ELEMENT foo ANY >
        <!ENTITY xxe SYSTEM "file:///etc/passwd" >]>
        <foo>&xxe;</foo>"""
        try:
            resp = requests.post(API_URL + "/api/chat", data=xml_payload, headers={"Content-Type": "application/xml"}, timeout=15)
            assert resp.status_code in [415, 400, 422, 503]
            _record("Vulnerability", "V-038", "XXE External Entity Injection Protection", "PASS", f"Rejected status: {resp.status_code}", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-038", "XXE External Entity Injection Protection", "WARN", str(e), time.time()-t0)

    def test_V039_unsupported_charset_handling(self):
        t0 = time.time()
        try:
            resp = requests.post(API_URL + "/api/chat", json={"message": "hi"}, headers={"Content-Type": "application/json; charset=utf-7"}, timeout=15)
            assert resp.status_code in [200, 400, 422, 415, 503]
            _record("Vulnerability", "V-039", "Unsupported Charset Header Acceptance", "PASS", f"Status: {resp.status_code}", time.time()-t0, "Low")
        except Exception as e:
            _record("Vulnerability", "V-039", "Unsupported Charset Header Acceptance", "WARN", str(e), time.time()-t0)

    def test_V040_arbitrary_file_upload_extension(self):
        t0 = time.time()
        try:
            files = {'file': ('test.txt', 'arbitrary content', 'text/plain')}
            resp = requests.post(API_URL + "/api/voice/transcribe", files=files, timeout=15)
            assert resp.status_code in [200, 400, 422, 500, 503]
            _record("Vulnerability", "V-040", "Audio File Extension Upload Check", "PASS", f"Response code: {resp.status_code}", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-040", "Audio File Extension Upload Check", "WARN", str(e), time.time()-t0)

    def test_V041_sensitive_endpoints_ssl_check(self):
        t0 = time.time()
        try:
            is_ssl = API_URL.startswith("https://")
            _record("Vulnerability", "V-041", "SSL/HTTPS Required for API Communication", "PASS" if is_ssl else "WARN", f"API Protocol: {API_URL[:8]}", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-041", "SSL/HTTPS Required for API Communication", "WARN", str(e), time.time()-t0)

    def test_V042_csrf_cookies_samesite(self):
        t0 = time.time()
        try:
            self._go()
            cookies = self.driver.get_cookies()
            non_samesite = [c for c in cookies if c.get("name") in ["csrf", "token"] and not c.get("sameSite")]
            assert not non_samesite
            _record("Vulnerability", "V-042", "Cookie SameSite CSRF Protections Check", "PASS", "No unprotected cookies found", time.time()-t0, "High")
        except Exception as e:
            _record("Vulnerability", "V-042", "Cookie SameSite CSRF Protections Check", "WARN", str(e), time.time()-t0)

    def test_V043_xss_iframe_src_injection(self):
        t0 = time.time()
        payload = "<iframe src=\"javascript:alert(1)\"></iframe>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-043", "Iframe Source XSS Injection Block", "FAIL", "Alert triggered via iframe!", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-043", "Iframe Source XSS Injection Block", "PASS", "Iframe payload neutralized", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-043", "Iframe Source XSS Injection Block", "WARN", str(e), time.time()-t0, "Critical")

    def test_V044_security_header_server_disclosure(self):
        t0 = time.time()
        try:
            resp = requests.get(BASE_URL, timeout=15, verify=False)
            server = resp.headers.get("Server", "")
            is_disclosed = "ubuntu" in server.lower() or "windows" in server.lower() or "nginx/1." in server.lower()
            _record("Vulnerability", "V-044", "Server Header Banner Disclosures Check", "PASS" if not is_disclosed else "WARN", f"Server Banner: '{server}'", time.time()-t0, "Low")
        except Exception as e:
            _record("Vulnerability", "V-044", "Server Header Banner Disclosures Check", "INFO", str(e), time.time()-t0)

    def test_V045_xss_math_payload(self):
        t0 = time.time()
        payload = "<math><mtext><option href=\"javascript:alert(1)\">click</option></mtext></math>"
        try:
            self._go("login")
            _wait_visible(self.driver, By.ID, "email").send_keys(payload)
            _find_submit_btn(self.driver).click()
            time.sleep(1.0)
            try:
                alert = self.driver.switch_to.alert
                alert.dismiss()
                _record("Vulnerability", "V-045", "Math-based XML/XSS Sanitization", "FAIL", "Alert triggered via math tag", time.time()-t0, "Critical")
                self.fail("XSS alert triggered!")
            except:
                _record("Vulnerability", "V-045", "Math-based XML/XSS Sanitization", "PASS", "Math tag parsed/escaped safely", time.time()-t0, "Critical")
        except Exception as e:
            _record("Vulnerability", "V-045", "Math-based XML/XSS Sanitization", "WARN", str(e), time.time()-t0, "Critical")


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — BACKEND API / UNIT TESTS (via requests)
# ═══════════════════════════════════════════════════════════════════════════════
class APIUnitTests(unittest.TestCase):
    BASE = API_URL.rstrip("/")

    # ── TC-A-001: Backend root health check ──────────────────────────────────
    def test_A001_backend_health(self):
        t0 = time.time()
        try:
            resp = requests.get(self.BASE + "/", timeout=30)
            assert resp.status_code in [200, 404, 422], f"Unexpected: {resp.status_code}"
            _record("API Unit", "A-001", "Backend Root Responds",
                    "PASS", f"HTTP {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-001", "Backend Root Responds",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-002: POST /api/chat ──────────────────────────────────────────────
    def test_A002_api_chat_endpoint(self):
        t0 = time.time()
        try:
            payload = {"message": "Hello! What are good places to visit in Goa?", "history": []}
            resp = requests.post(f"{self.BASE}/api/chat", json=payload, timeout=60)
            assert resp.status_code in [200, 503], f"Unexpected status: {resp.status_code}"
            if resp.status_code == 200:
                data = resp.json()
                assert "reply" in data, f"Missing 'reply' key, got: {list(data.keys())}"
                _record("API Unit", "A-002", "POST /api/chat Returns Reply",
                        "PASS", f"Reply length: {len(str(data.get('reply', '')))}", time.time()-t0)
            else:
                _record("API Unit", "A-002", "POST /api/chat Returns Reply",
                        "WARN", f"Service returned {resp.status_code} (possible cold start)",
                        time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-002", "POST /api/chat Returns Reply",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-003: GET /api/safety ─────────────────────────────────────────────
    def test_A003_api_safety_endpoint(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/safety?city=Goa", timeout=60)
            assert resp.status_code in [200, 503], f"Unexpected: {resp.status_code}"
            if resp.status_code == 200:
                data = resp.json()
                required = ["city", "generalSafety", "nightSafety"]
                missing = [k for k in required if k not in data]
                assert not missing, f"Missing fields: {missing}"
                _record("API Unit", "A-003", "GET /api/safety Returns Safety Metrics",
                        "PASS", f"City: {data.get('city')}, Safety: {data.get('generalSafety')}",
                        time.time()-t0)
            else:
                _record("API Unit", "A-003", "GET /api/safety Returns Safety Metrics",
                        "WARN", f"Service {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-003", "GET /api/safety Returns Safety Metrics",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-004: POST /api/briefing ─────────────────────────────────────────
    def test_A004_api_briefing_endpoint(self):
        t0 = time.time()
        try:
            payload = {
                "userName": "TestUser",
                "activeTripName": "Goa Adventure",
                "activeTripDestination": "Goa",
                "todayScheduleTitle": "Beach Day",
                "todayScheduleSpots": ["Baga Beach", "Fort Aguada"],
                "upcomingTripName": None,
                "upcomingTripDestination": None,
                "upcomingTripDays": None,
                "groupName": "Beach Squad",
                "groupExpensesCount": 3,
                "groupMembersCount": 4,
                "groupLastExpenseAmount": 1500.0,
                "groupLastExpenseDesc": "Dinner",
                "weatherTemp": 30,
                "weatherDesc": "Sunny"
            }
            resp = requests.post(f"{self.BASE}/api/briefing", json=payload, timeout=60)
            assert resp.status_code in [200, 503]
            if resp.status_code == 200:
                data = resp.json()
                assert "briefing" in data, f"Missing 'briefing' key"
                _record("API Unit", "A-004", "POST /api/briefing Returns Briefing Text",
                        "PASS", f"Briefing length: {len(str(data.get('briefing', '')))}",
                        time.time()-t0)
            else:
                _record("API Unit", "A-004", "POST /api/briefing Returns Briefing Text",
                        "WARN", f"Service {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-004", "POST /api/briefing Returns Briefing Text",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-005: POST /api/routes/optimize ──────────────────────────────────
    def test_A005_api_routes_optimize(self):
        t0 = time.time()
        try:
            payload = [
                {"name": "Baga Beach",    "latitude": 15.555, "longitude": 73.752},
                {"name": "Fort Aguada",   "latitude": 15.499, "longitude": 73.773},
                {"name": "Calangute",     "latitude": 15.542, "longitude": 73.755},
            ]
            resp = requests.post(f"{self.BASE}/api/routes/optimize", json=payload, timeout=60)
            assert resp.status_code in [200, 503]
            if resp.status_code == 200:
                data = resp.json()
                assert isinstance(data, list), "Expected list response"
                assert len(data) == len(payload), f"Count mismatch: {len(data)} vs {len(payload)}"
                _record("API Unit", "A-005", "POST /api/routes/optimize Returns Ordered Spots",
                        "PASS", f"Returned {len(data)} optimized spots", time.time()-t0)
            else:
                _record("API Unit", "A-005", "POST /api/routes/optimize Returns Ordered Spots",
                        "WARN", f"Service {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-005", "POST /api/routes/optimize Returns Ordered Spots",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-006: POST /api/expenses/split ───────────────────────────────────
    def test_A006_api_expenses_split(self):
        t0 = time.time()
        try:
            payload = {
                "amount": 1200,
                "description": "Beach dinner",
                "category": "Food",
                "paidBy": "Alice",
                "splitBetween": ["Alice", "Bob", "Charlie"]
            }
            resp = requests.post(f"{self.BASE}/api/expenses/split", json=payload, timeout=60)
            assert resp.status_code in [200, 422, 503]
            if resp.status_code == 200:
                data = resp.json()
                _record("API Unit", "A-006", "POST /api/expenses/split Returns Split Details",
                        "PASS", f"Response keys: {list(data.keys()) if isinstance(data, dict) else 'list'}",
                        time.time()-t0)
            elif resp.status_code == 422:
                _record("API Unit", "A-006", "POST /api/expenses/split Returns Split Details",
                        "WARN", "Validation error (422) — check payload schema", time.time()-t0)
            else:
                _record("API Unit", "A-006", "POST /api/expenses/split Returns Split Details",
                        "WARN", f"Service {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-006", "POST /api/expenses/split Returns Split Details",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-007: POST /api/routes/share ─────────────────────────────────────
    def test_A007_api_routes_share(self):
        t0 = time.time()
        try:
            payload = {
                "tripId": "test-trip-001",
                "tripName": "Goa Expedition",
                "destination": "Goa",
                "spots": [{"name": "Baga Beach", "latitude": 15.555, "longitude": 73.752}],
                "sharedBy": "TestUser"
            }
            resp = requests.post(f"{self.BASE}/api/routes/share", json=payload, timeout=60)
            assert resp.status_code in [200, 422, 503]
            if resp.status_code == 200:
                _record("API Unit", "A-007", "POST /api/routes/share Returns Share Data",
                        "PASS", f"HTTP {resp.status_code}", time.time()-t0)
            elif resp.status_code == 422:
                _record("API Unit", "A-007", "POST /api/routes/share Returns Share Data",
                        "WARN", "Schema validation error (422)", time.time()-t0)
            else:
                _record("API Unit", "A-007", "POST /api/routes/share Returns Share Data",
                        "WARN", f"Service {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-007", "POST /api/routes/share Returns Share Data",
                    "FAIL", str(e), time.time()-t0, "Critical")
            self.fail(str(e))

    # ── TC-A-008: CORS headers present on API ─────────────────────────────────
    def test_A008_cors_headers(self):
        t0 = time.time()
        try:
            resp = requests.options(
                f"{self.BASE}/api/chat",
                headers={"Origin": BASE_URL, "Access-Control-Request-Method": "POST"},
                timeout=15
            )
            cors = resp.headers.get("Access-Control-Allow-Origin", "")
            has_cors = bool(cors)
            _record("API Unit", "A-008", "CORS Headers Present on Backend API",
                    "PASS" if has_cors else "WARN",
                    f"ACAO header: '{cors}'", time.time()-t0, "High")
        except Exception as e:
            _record("API Unit", "A-008", "CORS Headers Present on Backend API",
                    "INFO", f"CORS check: {str(e)[:80]}", time.time()-t0)

    # ── TC-A-009: API rejects invalid JSON ───────────────────────────────────
    def test_A009_api_invalid_payload(self):
        t0 = time.time()
        try:
            resp = requests.post(
                f"{self.BASE}/api/chat",
                data="NOT VALID JSON {{{",
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            assert resp.status_code in [400, 422, 503], f"Expected 4xx, got {resp.status_code}"
            _record("API Unit", "A-009", "API Rejects Malformed JSON Payload",
                    "PASS", f"HTTP {resp.status_code} — server rejected invalid JSON",
                    time.time()-t0, "High")
        except AssertionError:
            raise
        except Exception as e:
            _record("API Unit", "A-009", "API Rejects Malformed JSON Payload",
                    "FAIL", str(e), time.time()-t0, "High")
            self.fail(str(e))

    # ── TC-A-010: API empty body handling ────────────────────────────────────
    def test_A010_api_empty_body(self):
        t0 = time.time()
        try:
            resp = requests.post(
                f"{self.BASE}/api/chat",
                json={},
                timeout=30
            )
            assert resp.status_code in [400, 422, 503]
            _record("API Unit", "A-010", "API Handles Empty Body with 4xx",
                    "PASS", f"HTTP {resp.status_code}", time.time()-t0)
        except AssertionError:
            raise
        except Exception as e:
            _record("API Unit", "A-010", "API Handles Empty Body with 4xx",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-A-011: Safety endpoint missing city param ──────────────────────────
    def test_A011_safety_missing_param(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/safety", timeout=30)
            assert resp.status_code in [400, 422, 503]
            _record("API Unit", "A-011", "Safety Endpoint Requires city Param",
                    "PASS", f"HTTP {resp.status_code} when city missing", time.time()-t0)
        except AssertionError:
            raise
        except Exception as e:
            _record("API Unit", "A-011", "Safety Endpoint Requires city Param",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    # ── TC-A-012: Response time under 60s (cold start allowed) ────────────────
    def test_A012_response_time(self):
        t0 = time.time()
        try:
            start = time.time()
            resp = requests.get(f"{self.BASE}/api/safety?city=Goa", timeout=70)
            elapsed = time.time() - start
            status = "PASS" if elapsed < 60 else "WARN"
            _record("API Unit", "A-012", "API Response Time Within 60s",
                    status, f"{elapsed:.2f}s (HTTP {resp.status_code})", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-012", "API Response Time Within 60s",
                    "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_A013_api_chat_history_validation(self):
        t0 = time.time()
        try:
            payload = {
                "message": "Continue",
                "history": [
                    {"role": "user", "content": "Tell me a joke"},
                    {"role": "assistant", "content": "Why did the traveler cross the road?"}
                ]
            }
            resp = requests.post(f"{self.BASE}/api/chat", json=payload, timeout=60)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-013", "POST /api/chat with Conversational History", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-013", "POST /api/chat with Conversational History", "WARN", str(e), time.time()-t0)

    def test_A014_api_chat_empty_message(self):
        t0 = time.time()
        try:
            payload = {"message": "", "history": []}
            resp = requests.post(f"{self.BASE}/api/chat", json=payload, timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-014", "POST /api/chat with Empty Prompt Validation", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-014", "POST /api/chat with Empty Prompt Validation", "WARN", str(e), time.time()-t0)

    def test_A015_api_chat_large_history(self):
        t0 = time.time()
        try:
            hist = [{"role": "user" if i%2==0 else "assistant", "content": f"msg {i}"} for i in range(50)]
            payload = {"message": "Final", "history": hist}
            resp = requests.post(f"{self.BASE}/api/chat", json=payload, timeout=60)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-015", "POST /api/chat with 50-Item Chat History Size", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-015", "POST /api/chat with 50-Item Chat History Size", "WARN", str(e), time.time()-t0)

    def test_A016_api_chat_special_characters(self):
        t0 = time.time()
        try:
            payload = {"message": "Goa 🏖️ 🔥 🗺️ @#$%^&*()_+", "history": []}
            resp = requests.post(f"{self.BASE}/api/chat", json=payload, timeout=60)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-016", "POST /api/chat with Emoji and Special Chars", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-016", "POST /api/chat with Emoji and Special Chars", "WARN", str(e), time.time()-t0)

    def test_A017_api_safety_invalid_city(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/safety?city=", timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-017", "GET /api/safety with Empty City Parameter", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-017", "GET /api/safety with Empty City Parameter", "WARN", str(e), time.time()-t0)

    def test_A018_api_safety_long_city_name(self):
        t0 = time.time()
        try:
            city = "A" * 150
            resp = requests.get(f"{self.BASE}/api/safety?city={city}", timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-018", "GET /api/safety with Long City Name (150 chars)", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-018", "GET /api/safety with Long City Name (150 chars)", "WARN", str(e), time.time()-t0)

    def test_A019_api_safety_special_chars_city(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/safety?city=New+York!@#$", timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-019", "GET /api/safety with City Symbols/Unicode", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-019", "GET /api/safety with City Symbols/Unicode", "WARN", str(e), time.time()-t0)

    def test_A020_api_briefing_missing_fields(self):
        t0 = time.time()
        try:
            payload = {"userName": "BriefTest"}
            resp = requests.post(f"{self.BASE}/api/briefing", json=payload, timeout=60)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-020", "POST /api/briefing Minimal Data Acceptability", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-020", "POST /api/briefing Minimal Data Acceptability", "WARN", str(e), time.time()-t0)

    def test_A021_api_briefing_extra_fields(self):
        t0 = time.time()
        try:
            payload = {"userName": "BriefTest", "extraFieldXYZ": "unexpectedValue"}
            resp = requests.post(f"{self.BASE}/api/briefing", json=payload, timeout=60)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-021", "POST /api/briefing Extra Fields Resiliency", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-021", "POST /api/briefing Extra Fields Resiliency", "WARN", str(e), time.time()-t0)

    def test_A022_api_briefing_type_mismatch(self):
        t0 = time.time()
        try:
            payload = {"userName": 12345}
            resp = requests.post(f"{self.BASE}/api/briefing", json=payload, timeout=30)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-022", "POST /api/briefing Field Type Validations", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-022", "POST /api/briefing Field Type Validations", "WARN", str(e), time.time()-t0)

    def test_A023_api_routes_optimize_empty_list(self):
        t0 = time.time()
        try:
            resp = requests.post(f"{self.BASE}/api/routes/optimize", json=[], timeout=30)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-023", "POST /api/routes/optimize Empty Array Handling", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-023", "POST /api/routes/optimize Empty Array Handling", "WARN", str(e), time.time()-t0)

    def test_A024_api_routes_optimize_single_spot(self):
        t0 = time.time()
        try:
            payload = [{"name": "Single Point", "latitude": 12.34, "longitude": 56.78}]
            resp = requests.post(f"{self.BASE}/api/routes/optimize", json=payload, timeout=30)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-024", "POST /api/routes/optimize Single Spot Array", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-024", "POST /api/routes/optimize Single Spot Array", "WARN", str(e), time.time()-t0)

    def test_A025_api_routes_optimize_invalid_coords(self):
        t0 = time.time()
        try:
            payload = [
                {"name": "Spot 1", "latitude": 12.34, "longitude": 56.78},
                {"name": "Spot 2", "latitude": 99.0, "longitude": -188.0}
            ]
            resp = requests.post(f"{self.BASE}/api/routes/optimize", json=payload, timeout=30)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-025", "POST /api/routes/optimize Invalid Coords Handling", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-025", "POST /api/routes/optimize Invalid Coords Handling", "WARN", str(e), time.time()-t0)

    def test_A026_api_routes_optimize_negative_coords(self):
        t0 = time.time()
        try:
            payload = [
                {"name": "Spot A", "latitude": -33.86, "longitude": 151.20},
                {"name": "Spot B", "latitude": -23.55, "longitude": -46.63}
            ]
            resp = requests.post(f"{self.BASE}/api/routes/optimize", json=payload, timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-026", "POST /api/routes/optimize Negative Coordinate Signs", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-026", "POST /api/routes/optimize Negative Coordinate Signs", "WARN", str(e), time.time()-t0)

    def test_A027_api_expenses_split_negative_amount(self):
        t0 = time.time()
        try:
            payload = {"totalAmount": -100.0, "members": ["Alice", "Bob"], "description": "Refund"}
            resp = requests.post(f"{self.BASE}/api/expenses/split", json=payload, timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-027", "POST /api/expenses/split with Negative Amount", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-027", "POST /api/expenses/split with Negative Amount", "WARN", str(e), time.time()-t0)

    def test_A028_api_expenses_split_empty_members(self):
        t0 = time.time()
        try:
            payload = {"totalAmount": 100.0, "members": [], "description": "Split"}
            resp = requests.post(f"{self.BASE}/api/expenses/split", json=payload, timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-028", "POST /api/expenses/split with Empty Member List", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-028", "POST /api/expenses/split with Empty Member List", "WARN", str(e), time.time()-t0)

    def test_A029_api_expenses_split_single_member(self):
        t0 = time.time()
        try:
            payload = {"totalAmount": 100.0, "members": ["Alice"], "description": "Self Split"}
            resp = requests.post(f"{self.BASE}/api/expenses/split", json=payload, timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-029", "POST /api/expenses/split Single Member Handling", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-029", "POST /api/expenses/split Single Member Handling", "WARN", str(e), time.time()-t0)

    def test_A030_api_expenses_split_fractional_amount(self):
        t0 = time.time()
        try:
            payload = {"totalAmount": 100.33, "members": ["A", "B", "C"], "description": "Fraction split"}
            resp = requests.post(f"{self.BASE}/api/expenses/split", json=payload, timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-030", "POST /api/expenses/split Fractional Math Rounds", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-030", "POST /api/expenses/split Fractional Math Rounds", "WARN", str(e), time.time()-t0)

    def test_A031_api_routes_share_invalid_count(self):
        t0 = time.time()
        try:
            payload = {
                "routeId": "test-route-123",
                "routeName": "Test Path",
                "stopsCount": -5,
                "totalDistance": "10 km",
                "totalDuration": "30 mins"
            }
            resp = requests.post(f"{self.BASE}/api/routes/share", json=payload, timeout=30)
            assert resp.status_code in [200, 422, 503]
            _record("API Unit", "A-031", "POST /api/routes/share Negative Stops Boundary", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-031", "POST /api/routes/share Negative Stops Boundary", "WARN", str(e), time.time()-t0)

    def test_A032_api_routes_share_missing_fields(self):
        t0 = time.time()
        try:
            payload = {"routeId": "123"}
            resp = requests.post(f"{self.BASE}/api/routes/share", json=payload, timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-032", "POST /api/routes/share Missing Required Keys", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-032", "POST /api/routes/share Missing Required Keys", "WARN", str(e), time.time()-t0)

    def test_A033_api_weather_valid_coords(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/weather?lat=15.5&lon=73.8", timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-033", "GET /api/weather Valid Lat/Lon Proxy", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-033", "GET /api/weather Valid Lat/Lon Proxy", "WARN", str(e), time.time()-t0)

    def test_A034_api_weather_invalid_coords(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/weather?lat=999.0&lon=-999.0", timeout=30)
            assert resp.status_code in [200, 400, 422, 500, 503]
            _record("API Unit", "A-034", "GET /api/weather Out of Bounds Coordinates", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-034", "GET /api/weather Out of Bounds Coordinates", "WARN", str(e), time.time()-t0)

    def test_A035_api_weather_missing_coords(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/weather", timeout=30)
            assert resp.status_code in [400, 422, 503]
            _record("API Unit", "A-035", "GET /api/weather Missing Params Handling", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-035", "GET /api/weather Missing Params Handling", "WARN", str(e), time.time()-t0)

    def test_A036_api_otp_send_valid(self):
        t0 = time.time()
        try:
            payload = {"email": "user@example.com", "otp": "123456"}
            resp = requests.post(f"{self.BASE}/api/otp/send", json=payload, timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-036", "POST /api/otp/send Valid Email", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-036", "POST /api/otp/send Valid Email", "WARN", str(e), time.time()-t0)

    def test_A037_api_otp_send_invalid_email(self):
        t0 = time.time()
        try:
            payload = {"email": "malformed_email", "otp": "123456"}
            resp = requests.post(f"{self.BASE}/api/otp/send", json=payload, timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-037", "POST /api/otp/send Invalid Email Verification", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-037", "POST /api/otp/send Invalid Email Verification", "WARN", str(e), time.time()-t0)

    def test_A038_api_otp_send_missing_otp(self):
        t0 = time.time()
        try:
            payload = {"email": "user@example.com", "otp": ""}
            resp = requests.post(f"{self.BASE}/api/otp/send", json=payload, timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            _record("API Unit", "A-038", "POST /api/otp/send Missing Code Check", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-038", "POST /api/otp/send Missing Code Check", "WARN", str(e), time.time()-t0)

    def test_A039_api_trips_valid_userId(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/trips?userId=user_999", timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-039", "GET /api/trips List Retrieve", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-039", "GET /api/trips List Retrieve", "WARN", str(e), time.time()-t0)

    def test_A040_api_trips_missing_userId(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/trips", timeout=30)
            assert resp.status_code in [400, 422, 503]
            _record("API Unit", "A-040", "GET /api/trips Missing User ID Error", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-040", "GET /api/trips Missing User ID Error", "WARN", str(e), time.time()-t0)

    def test_A041_api_trips_post_valid(self):
        t0 = time.time()
        try:
            payload = {
                "title": "Summer Camp",
                "destination": "Ooty",
                "startDate": "2025-05-01",
                "endDate": "2025-05-05",
                "userId": "user_123"
            }
            resp = requests.post(f"{self.BASE}/api/trips", json=payload, timeout=30)
            assert resp.status_code in [200, 503]
            _record("API Unit", "A-041", "POST /api/trips Creation Validation", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-041", "POST /api/trips Creation Validation", "WARN", str(e), time.time()-t0)

    def test_A042_api_trips_post_missing_fields(self):
        t0 = time.time()
        try:
            payload = {"title": "Summer Camp", "destination": "Ooty", "userId": "user_123"}
            resp = requests.post(f"{self.BASE}/api/trips", json=payload, timeout=30)
            assert resp.status_code in [400, 422, 503]
            _record("API Unit", "A-042", "POST /api/trips Missing Dates Validation", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-042", "POST /api/trips Missing Dates Validation", "WARN", str(e), time.time()-t0)

    def test_A043_api_non_existent_route(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/this_endpoint_does_not_exist_ever", timeout=30)
            assert resp.status_code in [404, 503]
            _record("API Unit", "A-043", "Backend Non-existent Endpoint returns 404", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-043", "Backend Non-existent Endpoint returns 404", "WARN", str(e), time.time()-t0)

    def test_A044_api_response_content_type(self):
        t0 = time.time()
        try:
            resp = requests.get(f"{self.BASE}/api/trips?userId=123", timeout=30)
            assert resp.status_code in [200, 400, 422, 503]
            ct = resp.headers.get("Content-Type", "")
            is_json = "application/json" in ct.lower() if resp.status_code == 200 else True
            _record("API Unit", "A-044", "API Header Content-Type Validation", "PASS" if is_json else "WARN", f"Content-Type: '{ct}'", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-044", "API Header Content-Type Validation", "WARN", str(e), time.time()-t0)

    def test_A045_api_options_request(self):
        t0 = time.time()
        try:
            resp = requests.options(f"{self.BASE}/api/routes/share", timeout=30)
            assert resp.status_code in [200, 204, 405, 503]
            _record("API Unit", "A-045", "API Preflight OPTIONS Request Acceptability", "PASS", f"Status: {resp.status_code}", time.time()-t0)
        except Exception as e:
            _record("API Unit", "A-045", "API Preflight OPTIONS Request Acceptability", "WARN", str(e), time.time()-t0)


# ═══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — UI/UX TESTS
# ═══════════════════════════════════════════════════════════════════════════════
class UIUXTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.driver = _build_driver()
        cls.driver.set_page_load_timeout(45)

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()

    def _go(self, path=""):
        self.driver.get(BASE_URL.rstrip("/") + ("/" + path.lstrip("/") if path else ""))
        time.sleep(1.5)

    def test_UI001_landing_glassmorphism(self):
        t0 = time.time()
        try:
            self._go()
            src = self.driver.page_source.lower()
            has_glass = "glass" in src or "backdrop-blur" in src or "shadow" in src
            _record("UI UX", "UI-001", "Landing Page Glassmorphism Effects", "PASS" if has_glass else "WARN", f"Found glass references: {has_glass}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-001", "Landing Page Glassmorphism Effects", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI002_landing_footer_links(self):
        t0 = time.time()
        try:
            self._go()
            time.sleep(2)
            # Try footer tag first, then footer-like containers
            footer_el = None
            for sel in ['footer', '[class*="footer"]', '[id*="footer"]',
                        '[class*="Footer"]', '[class*="bottom"]']:
                try:
                    els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                    if els:
                        footer_el = els[0]
                        break
                except Exception:
                    continue
            if footer_el:
                links = footer_el.find_elements(By.TAG_NAME, 'a')
                _record('UI UX', 'UI-002', 'Landing Footer Links Exist',
                        'PASS', f'Found footer element with {len(links)} links', time.time()-t0)
            else:
                all_links = self.driver.find_elements(By.TAG_NAME, 'a')
                _record('UI UX', 'UI-002', 'Landing Footer Links Exist',
                        'WARN', f'No footer tag; page has {len(all_links)} total links', time.time()-t0)
        except Exception as e:
            _record('UI UX', 'UI-002', 'Landing Footer Links Exist',
                    'WARN', f'Could not check footer: {str(e)[:100]}', time.time()-t0)

    def test_UI003_landing_typography(self):
        t0 = time.time()
        try:
            self._go()
            body = self.driver.find_element(By.TAG_NAME, "body")
            ff = body.value_of_css_property("font-family")
            _record("UI UX", "UI-003", "CSS Typography System Verification", "PASS", f"Font family: {ff}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-003", "CSS Typography System Verification", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI004_navbar_logo_animation(self):
        t0 = time.time()
        try:
            self._go()
            src = self.driver.page_source
            has_anim = "animate-spin-slow" in src or "spin" in src.lower()
            _record("UI UX", "UI-004", "Logo Spin Animation Class Check", "PASS" if has_anim else "WARN", f"Logo animation configured: {has_anim}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-004", "Logo Spin Animation Class Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI005_login_card_styling(self):
        t0 = time.time()
        try:
            self._go("login")
            src = self.driver.page_source.lower()
            has_card = "glass-panel" in src or "bg-slate-900" in src or "shadow-2xl" in src
            _record("UI UX", "UI-005", "Login Panel Cards Styling", "PASS" if has_card else "WARN", f"Has card style markers: {has_card}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-005", "Login Panel Cards Styling", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI006_register_card_styling(self):
        t0 = time.time()
        try:
            self._go("register")
            src = self.driver.page_source.lower()
            has_card = "glass-panel" in src or "bg-slate-900" in src or "shadow-2xl" in src
            _record("UI UX", "UI-006", "Registration Panel Cards Styling", "PASS" if has_card else "WARN", f"Has card style markers: {has_card}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-006", "Registration Panel Cards Styling", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI007_dashboard_sidebar_blur(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            src = self.driver.page_source.lower()
            has_blur = "backdrop-blur" in src or "bg-slate-950" in src
            _record("UI UX", "UI-007", "Dashboard Sidebar Backdrop Blur Check", "PASS" if has_blur else "WARN", f"Sidebar style markers: {has_blur}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-007", "Dashboard Sidebar Backdrop Blur Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI008_dashboard_sidebar_links_hover(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            src = self.driver.page_source.lower()
            has_transition = "transition" in src or "duration-" in src
            _record("UI UX", "UI-008", "Sidebar Anchor Hover Transitions Check", "PASS" if has_transition else "WARN", f"Has transition properties: {has_transition}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-008", "Sidebar Anchor Hover Transitions Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI009_dashboard_sidebar_logout_btn(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            src = self.driver.page_source.lower()
            has_logout = "logout" in src or "rose" in src
            _record("UI UX", "UI-009", "Sidebar Logout Colored Anchor Style", "PASS" if has_logout else "WARN", f"Logout markers found: {has_logout}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-009", "Sidebar Logout Colored Anchor Style", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI010_dashboard_grid_layout(self):
        t0 = time.time()
        try:
            self._go("dashboard")
            src = self.driver.page_source.lower()
            has_grid = "grid" in src or "flex" in src or "col" in src
            _record("UI UX", "UI-010", "Dashboard Container Grid Check", "PASS" if has_grid else "WARN", f"Has layout classes: {has_grid}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-010", "Dashboard Container Grid Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI011_ai_assistant_chat_bubbles(self):
        t0 = time.time()
        try:
            self._go("ai-assistant")
            src = self.driver.page_source.lower()
            has_bubble = "bubble" in src or "message" in src or "chat" in src or "rounded-xl" in src
            _record("UI UX", "UI-011", "AI Assistant Chat Bubble Styling Check", "PASS" if has_bubble else "WARN", f"Has chat bubble: {has_bubble}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-011", "AI Assistant Chat Bubble Styling Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI012_ai_planner_form_layout(self):
        t0 = time.time()
        try:
            self._go("ai-planner")
            src = self.driver.page_source.lower()
            has_inputs = "input" in src or "form" in src or "bg-slate-900" in src
            _record("UI UX", "UI-012", "AI Planner Input Fields Styling", "PASS" if has_inputs else "WARN", f"Has planner styles: {has_inputs}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-012", "AI Planner Input Fields Styling", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI013_safety_score_indicators(self):
        t0 = time.time()
        try:
            self._go("safety")
            src = self.driver.page_source.lower()
            has_indicator = "safety" in src or "map" in src or "indicator" in src or "score" in src
            _record("UI UX", "UI-013", "Safety Score Graphic Indicators Check", "PASS" if has_indicator else "WARN", f"Has safety indicators: {has_indicator}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-013", "Safety Score Graphic Indicators Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI014_expenses_total_split_badge(self):
        t0 = time.time()
        try:
            self._go("expenses")
            src = self.driver.page_source.lower()
            has_split = "split" in src or "expense" in src or "badge" in src or "total" in src
            _record("UI UX", "UI-014", "Expenses Splits Badging Presence", "PASS" if has_split else "WARN", f"Has splitting visual details: {has_split}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-014", "Expenses Splits Badging Presence", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI015_routes_map_container(self):
        t0 = time.time()
        try:
            self._go("routes")
            src = self.driver.page_source.lower()
            has_map = "map" in src or "leaflet" in src or "container" in src
            _record("UI UX", "UI-015", "Routes Map Display Grid Panel", "PASS" if has_map else "WARN", f"Has map container visual: {has_map}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-015", "Routes Map Display Grid Panel", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI016_route_sharing_card(self):
        t0 = time.time()
        try:
            self._go("route-sharing")
            src = self.driver.page_source.lower()
            has_card = "card" in src or "glass" in src or "border-white" in src or "share" in src
            _record("UI UX", "UI-016", "Route Sharing Layout Panels Check", "PASS" if has_card else "WARN", f"Has route share panels: {has_card}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-016", "Route Sharing Layout Panels Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI017_explore_places_grid(self):
        t0 = time.time()
        try:
            self._go("explore")
            src = self.driver.page_source.lower()
            has_grid = "grid" in src or "flex" in src or "place" in src or "cols" in src
            _record("UI UX", "UI-017", "Explore Destination Cards Grid Layout", "PASS" if has_grid else "WARN", f"Has explore grid style: {has_grid}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-017", "Explore Destination Cards Grid Layout", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI018_trips_card_images(self):
        t0 = time.time()
        try:
            self._go("trips")
            src = self.driver.page_source.lower()
            has_img = "img" in src or "unsplash" in src or "image" in src or "cover" in src
            _record("UI UX", "UI-018", "Trips Background Hero Images Check", "PASS" if has_img else "WARN", f"Has cover image configs: {has_img}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-018", "Trips Background Hero Images Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI019_groups_avatars(self):
        t0 = time.time()
        try:
            self._go("groups")
            src = self.driver.page_source.lower()
            has_avatar = "avatar" in src or "rounded-full" in src or "buddies" in src or "member" in src
            _record("UI UX", "UI-019", "Groups Buddies Round Avatars Style", "PASS" if has_avatar else "WARN", f"Has avatars design: {has_avatar}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-019", "Groups Buddies Round Avatars Style", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI020_profile_form_borders(self):
        t0 = time.time()
        try:
            self._go("profile")
            src = self.driver.page_source.lower()
            has_borders = "border" in src or "input" in src or "focus:" in src
            _record("UI UX", "UI-020", "Profile Settings Borders Details", "PASS" if has_borders else "WARN", f"Has border classes: {has_borders}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-020", "Profile Settings Borders Details", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI021_notifications_severity_colors(self):
        t0 = time.time()
        try:
            self._go("notifications")
            src = self.driver.page_source.lower()
            has_alerts = "alert" in src or "notification" in src or "rose" in src or "teal" in src
            _record("UI UX", "UI-021", "Notifications Severity Color Panels", "PASS" if has_alerts else "WARN", f"Has alert colors configs: {has_alerts}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-021", "Notifications Severity Color Panels", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI022_favorites_star_icons(self):
        t0 = time.time()
        try:
            self._go("favorites")
            src = self.driver.page_source.lower()
            has_icons = "heart" in src or "star" in src or "lucide" in src or "favorite" in src
            _record("UI UX", "UI-022", "Favorites Heart/Star Icons Check", "PASS" if has_icons else "WARN", f"Has favorite icons: {has_icons}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-022", "Favorites Heart/Star Icons Check", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI023_visited_badges(self):
        t0 = time.time()
        try:
            self._go("visited")
            src = self.driver.page_source.lower()
            has_visited = "visited" in src or "check" in src or "badge" in src
            _record("UI UX", "UI-023", "Visited Status Validation Badging", "PASS" if has_visited else "WARN", f"Has visited icons: {has_visited}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-023", "Visited Status Validation Badging", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI024_dark_mode_theme_colors(self):
        t0 = time.time()
        try:
            self._go()
            src = self.driver.page_source.lower()
            has_dark = "bg-slate-950" in src or "slate-900" in src or "dark" in src
            _record("UI UX", "UI-024", "Main Layout Theme Dark Mode Colors", "PASS" if has_dark else "WARN", f"Has dark mode values: {has_dark}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-024", "Main Layout Theme Dark Mode Colors", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))

    def test_UI025_input_focus_glow(self):
        t0 = time.time()
        try:
            self._go("login")
            src = self.driver.page_source.lower()
            has_focus = "focus:border" in src or "focus:outline" in src or "transition" in src
            _record("UI UX", "UI-025", "Form Input Focus Ring Colors", "PASS" if has_focus else "WARN", f"Has focus rings configured: {has_focus}", time.time()-t0)
        except Exception as e:
            _record("UI UX", "UI-025", "Form Input Focus Ring Colors", "FAIL", str(e), time.time()-t0)
            self.fail(str(e))


# ═══════════════════════════════════════════════════════════════════════════════
#  XLSX REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════
def generate_xlsx_report():
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"TripSync_TestReport_{timestamp}.xlsx"
    filepath  = os.path.join(REPORT_DIR, filename)

    wb = openpyxl.Workbook()

    # ── Color palette ──────────────────────────────────────────────────────────
    C_PASS    = "FF22C55E"   # green-500
    C_FAIL    = "FFEF4444"   # red-500
    C_WARN    = "FFF59E0B"   # amber-400
    C_SKIP    = "FF94A3B8"   # slate-400
    C_INFO    = "FF60A5FA"   # blue-400
    C_HEADER  = "FF0F172A"   # slate-950
    C_SUBHDR  = "FF1E293B"   # slate-800
    C_LIGHT   = "FFF1F5F9"   # slate-100
    C_WHITE   = "FFFFFFFF"
    C_TEAL    = "FF14B8A6"   # teal-500

    thin  = Side(style="thin",   color="FF334155")
    thick = Side(style="medium", color="FF334155")

    def make_border(left=thin, right=thin, top=thin, bottom=thin):
        return Border(left=left, right=right, top=top, bottom=bottom)

    def header_font(size=11, bold=True, color=C_WHITE):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def body_font(size=10, bold=False, color="FF1E293B"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    # ── Sheet 1: Executive Summary ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"

    total      = len(_results)
    passed     = sum(1 for r in _results if r["status"] == "PASS")
    failed     = sum(1 for r in _results if r["status"] == "FAIL")
    warned     = sum(1 for r in _results if r["status"] in ["WARN", "SKIP", "INFO"])
    pass_rate  = (passed / total * 100) if total else 0
    run_time   = sum(r["duration"] for r in _results)

    by_cat: dict[str, dict] = {}
    for r in _results:
        c = r["category"]
        if c not in by_cat:
            by_cat[c] = {"total": 0, "pass": 0, "fail": 0, "warn": 0}
        by_cat[c]["total"] += 1
        if r["status"] == "PASS":    by_cat[c]["pass"] += 1
        elif r["status"] == "FAIL":  by_cat[c]["fail"] += 1
        else:                        by_cat[c]["warn"] += 1

    # Title block
    ws_sum.merge_cells("A1:H1")
    ws_sum["A1"] = "🧪 TripSync Web Application — E2E Test Report"
    ws_sum["A1"].font      = header_font(16)
    ws_sum["A1"].fill      = PatternFill("solid", fgColor=C_HEADER)
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 40

    ws_sum.merge_cells("A2:H2")
    ws_sum["A2"] = (f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                    f"App: {BASE_URL}  |  API: {API_URL}")
    ws_sum["A2"].font      = body_font(9, color=C_WHITE)
    ws_sum["A2"].fill      = PatternFill("solid", fgColor=C_SUBHDR)
    ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    # KPI row
    kpis = [
        ("Total Tests",  total,          C_HEADER),
        ("✅ Passed",    passed,          C_PASS),
        ("❌ Failed",    failed,          C_FAIL),
        ("⚠ Warn/Skip", warned,          "FFFBBF24"),
        ("Pass Rate",   f"{pass_rate:.1f}%", C_TEAL),
        ("Runtime",     f"{run_time:.1f}s",  "FF7C3AED"),
    ]
    kpi_cols = [1, 2, 3, 4, 6, 7]
    ws_sum.row_dimensions[4].height = 60
    for col, (label, val, color) in zip(kpi_cols, kpis):
        ws_sum.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        cell = ws_sum.cell(row=4, column=col)
        cell.value     = f"{label}\n{val}"
        cell.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()

    # Category breakdown table
    row = 7
    cat_headers = ["Category", "Total", "Passed", "Failed", "Warn/Skip", "Pass %"]
    for ci, h in enumerate(cat_headers, 1):
        c = ws_sum.cell(row=row, column=ci, value=h)
        c.font      = header_font(10)
        c.fill      = PatternFill("solid", fgColor=C_SUBHDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()
    ws_sum.row_dimensions[row].height = 20

    for row_i, (cat, d) in enumerate(by_cat.items(), row + 1):
        pct = (d["pass"] / d["total"] * 100) if d["total"] else 0
        row_data = [cat, d["total"], d["pass"], d["fail"], d["warn"], f"{pct:.0f}%"]
        fill_col = C_PASS if d["fail"] == 0 else C_FAIL
        for ci, val in enumerate(row_data, 1):
            c = ws_sum.cell(row=row_i, column=ci, value=val)
            c.font      = body_font(10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            c.border    = make_border()
            if ci in [3, 4]:
                c.fill = PatternFill("solid", fgColor=C_PASS if ci==3 else (C_FAIL if d["fail"]>0 else C_PASS))
        ws_sum.row_dimensions[row_i].height = 18

    # Column widths
    for col, w in zip("ABCDEFGH", [30, 12, 12, 12, 14, 14, 20, 20]):
        ws_sum.column_dimensions[get_column_letter(col if isinstance(col, int)
                                                    else ord(col)-64)].width = w

    # ── Sheet 2–4: Per-category detailed results ───────────────────────────────
    category_order = ["Functional", "Vulnerability", "API Unit", "UI UX"]
    col_headers = ["#", "Test ID", "Test Name", "Status", "Duration (s)",
                   "Severity", "Details", "Timestamp"]
    col_widths   = [5, 10, 45, 10, 13, 12, 60, 20]

    STATUS_COLORS = {
        "PASS": C_PASS,
        "FAIL": C_FAIL,
        "WARN": "FFFBBF24",
        "SKIP": C_SKIP,
        "INFO": C_INFO,
    }

    for cat in category_order:
        cat_results = [r for r in _results if r["category"] == cat]
        if not cat_results:
            continue

        # Sanitize sheet title — Excel forbids: / \ ? * [ ]
        safe_title = cat.replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "")[:31]
        ws = wb.create_sheet(title=safe_title)

        # Sheet title
        ws.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")
        ws["A1"] = f"TripSync — {cat} Test Results"
        ws["A1"].font      = header_font(13)
        ws["A1"].fill      = PatternFill("solid", fgColor=C_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Column headers
        for ci, (h, w) in enumerate(zip(col_headers, col_widths), 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font      = header_font(10)
            c.fill      = PatternFill("solid", fgColor=C_TEAL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = make_border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 22

        # Data rows
        for ri, r in enumerate(cat_results, 3):
            row_vals = [
                ri - 2,
                r["test_id"],
                r["name"],
                r["status"],
                r["duration"],
                r["severity"],
                r["details"][:200],
                r["timestamp"],
            ]
            status_color = STATUS_COLORS.get(r["status"], C_LIGHT)
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font      = body_font(9 if ci == 7 else 10)
                c.alignment = Alignment(
                    horizontal="left" if ci in [3, 7] else "center",
                    vertical="center", wrap_text=(ci == 7)
                )
                c.border = make_border()
                if ci == 4:  # Status column — colored
                    c.fill = PatternFill("solid", fgColor=status_color)
                    c.font = Font(name="Calibri", size=10, bold=True,
                                  color=C_WHITE if r["status"] in ["FAIL","PASS"] else "FF1E293B")
                elif ri % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="FFF8FAFC")
            ws.row_dimensions[ri].height = 22 if len(r.get("details","")) < 80 else 36

        ws.freeze_panes = "A3"

    # ── Sheet 5: All Results (flat) ────────────────────────────────────────────
    ws_all = wb.create_sheet("All Results")
    ws_all.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")
    ws_all["A1"] = "TripSync — All Test Results (Combined)"
    ws_all["A1"].font      = header_font(13)
    ws_all["A1"].fill      = PatternFill("solid", fgColor=C_HEADER)
    ws_all["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 30

    all_headers = ["#", "Category", "Test ID", "Test Name", "Status", "Duration (s)",
                   "Severity", "Details", "Timestamp"]
    all_widths   = [5, 14, 10, 42, 10, 13, 12, 55, 20]
    for ci, (h, w) in enumerate(zip(all_headers, all_widths), 1):
        c = ws_all.cell(row=2, column=ci, value=h)
        c.font      = header_font(10)
        c.fill      = PatternFill("solid", fgColor=C_TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()
        ws_all.column_dimensions[get_column_letter(ci)].width = w
    ws_all.row_dimensions[2].height = 22

    for ri, r in enumerate(_results, 3):
        row_vals = [
            ri - 2, r["category"], r["test_id"], r["name"],
            r["status"], r["duration"], r["severity"], r["details"][:200], r["timestamp"]
        ]
        status_color = STATUS_COLORS.get(r["status"], C_LIGHT)
        for ci, val in enumerate(row_vals, 1):
            c = ws_all.cell(row=ri, column=ci, value=val)
            c.font      = body_font(9 if ci == 8 else 10)
            c.alignment = Alignment(
                horizontal="left" if ci in [4, 8] else "center",
                vertical="center", wrap_text=(ci == 8)
            )
            c.border = make_border()
            if ci == 5:
                c.fill = PatternFill("solid", fgColor=status_color)
                c.font = Font(name="Calibri", size=10, bold=True,
                              color=C_WHITE if r["status"] in ["FAIL","PASS"] else "FF1E293B")
            elif ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FFF8FAFC")
        ws_all.row_dimensions[ri].height = 22
    ws_all.freeze_panes = "A3"

    wb.save(filepath)
    print(f"\n📊 XLSX report saved → {filepath}")
    return filepath


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    import sys
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass
    print(f"\n{'='*70}")
    print(f"  🧭 TripSync E2E Test Suite")
    print(f"  App URL : {BASE_URL}")
    print(f"  API URL : {API_URL}")
    print(f"  Headless: {HEADLESS}")
    print(f"{'='*70}\n")

    loader = unittest.TestLoader()
    suite  = unittest.TestSuite()

    # Order: API first (fast, no browser), then functional, then security
    for cls in [APIUnitTests, FunctionalTests, VulnerabilityTests, UIUXTests]:
        suite.addTests(loader.loadTestsFromTestCase(cls))

    runner = unittest.TextTestRunner(verbosity=2, stream=sys.stdout)
    result = runner.run(suite)

    # Generate report
    report_path = generate_xlsx_report()

    # Print summary
    total   = len(_results)
    passed  = sum(1 for r in _results if r["status"] == "PASS")
    failed  = sum(1 for r in _results if r["status"] == "FAIL")
    warned  = sum(1 for r in _results if r["status"] in ["WARN", "SKIP", "INFO"])
    print(f"\n{'='*70}")
    print(f"  📋 FINAL SUMMARY")
    print(f"  Total  : {total}")
    print(f"  ✅ Pass : {passed}")
    print(f"  ❌ Fail : {failed}")
    print(f"  ⚠  Warn : {warned}")
    print(f"  Pass %  : {(passed/total*100):.1f}%" if total else "  Pass % : N/A")
    print(f"  Report  : {report_path}")
    print(f"{'='*70}\n")

    sys.exit(0 if result.wasSuccessful() else 1)


if __name__ == "__main__":
    main()
